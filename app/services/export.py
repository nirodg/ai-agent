"""CSV / XLSX / Markdown export helpers."""

import csv
import io

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill


CONF_FILL = {
    "high":   PatternFill("solid", fgColor="C6EFCE"),
    "medium": PatternFill("solid", fgColor="FFEB9C"),
    "low":    PatternFill("solid", fgColor="FFC7CE"),
}


def profile_to_markdown(p: dict) -> str:
    conf = p.get("confidence", {})
    bt   = {"high":"[HIGH]","medium":"[MEDIUM]","low":"[LOW]"}
    def b(f): return bt.get((conf.get(f) or "low").lower(),"[LOW]")
    md  = f"### 🏢 {p['company_name']}\n\n"
    md += f"**Core Product** {b('core_product')}\n{p['core_product']}\n\n"
    md += f"**Recent News** {b('recent_news')}\n{p['recent_news']}\n\n"
    md += f"### 🎯 Pain Points {b('pain_points')}\n\n"
    for pt in p["pain_points"]: md += f"- {pt}\n"
    md += f"\n### 💡 Pitch Angle {b('pitch_angle')}\n\n{p['pitch_angle']}\n"
    fi = p.get("funding_info",{})
    if fi:
        md += f"\n### 💰 Funding\n- Raised: {fi.get('total_raised','?')}\n"
        md += f"- Last round: {fi.get('last_round','?')}\n"
    js = p.get("job_signals",{})
    if js and js.get("hiring_themes","Unknown") != "Unknown":
        md += f"\n### 👥 Hiring\n{js.get('hiring_themes','')}\n"
    ts = p.get("tech_stack",{})
    if ts and ts.get("tools_identified"):
        md += f"\n### 🛠 Stack\n{', '.join(ts['tools_identified'])}\n"
    return md


def profiles_to_csv(profiles: list) -> bytes:
    output = io.StringIO()
    fieldnames = [
        "company_name","core_product","recent_news","pain_points","pitch_angle",
        "conf_core","conf_news","conf_pain","conf_pitch",
        "funding_raised","funding_round","funding_investors","funding_revenue","conf_funding",
        "jobs_roles","jobs_themes","jobs_headcount","jobs_pitch",
        "stack_tools","stack_summary","stack_pitch",
        "competitors","search_depth","created_at",
    ]
    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()
    for p in profiles:
        conf = p.get("confidence",{})
        fi   = p.get("funding_info",{})
        js   = p.get("job_signals",{})
        ts   = p.get("tech_stack",{})
        writer.writerow({
            "company_name":      p["company_name"],
            "core_product":      p["core_product"],
            "recent_news":       p["recent_news"],
            "pain_points":       "; ".join(p["pain_points"]),
            "pitch_angle":       p["pitch_angle"],
            "conf_core":         conf.get("core_product",""),
            "conf_news":         conf.get("recent_news",""),
            "conf_pain":         conf.get("pain_points",""),
            "conf_pitch":        conf.get("pitch_angle",""),
            "funding_raised":    fi.get("total_raised",""),
            "funding_round":     fi.get("last_round",""),
            "funding_investors": fi.get("key_investors",""),
            "funding_revenue":   fi.get("revenue_signals",""),
            "conf_funding":      conf.get("funding_info",""),
            "jobs_roles":        "; ".join(js.get("open_roles",[])),
            "jobs_themes":       js.get("hiring_themes",""),
            "jobs_headcount":    js.get("headcount_signal",""),
            "jobs_pitch":        js.get("pitch_implication",""),
            "stack_tools":       "; ".join(ts.get("tools_identified",[])),
            "stack_summary":     ts.get("stack_summary",""),
            "stack_pitch":       ts.get("pitch_implication",""),
            "competitors":       "; ".join(c.get("company_name","") for c in p.get("competitors",[])),
            "search_depth":      p.get("search_depth",""),
            "created_at":        p.get("created_at",""),
        })
    return output.getvalue().encode("utf-8")


def profiles_to_xlsx(profiles: list) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Enriched Leads"
    headers = [
        "Company","Core Product","Recent News","Pain Points","Pitch Angle",
        "Conf:Core","Conf:News","Conf:Pain","Conf:Pitch",
        "Funding Raised","Last Round","Investors","Revenue Signals","Conf:Funding",
        "Open Roles","Hiring Themes","Headcount","Jobs Pitch",
        "Tech Tools","Stack Summary","Stack Pitch",
        "Competitors","Depth","Researched At",
    ]
    hf = PatternFill("solid", fgColor="1F4E79")
    hfont = Font(bold=True, color="FFFFFF")
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill, c.font, c.alignment = hf, hfont, Alignment(wrap_text=True)

    for ri, p in enumerate(profiles, 2):
        conf = p.get("confidence",{})
        fi   = p.get("funding_info",{})
        js   = p.get("job_signals",{})
        ts   = p.get("tech_stack",{})
        vals = [
            p["company_name"], p["core_product"], p["recent_news"],
            "; ".join(p["pain_points"]), p["pitch_angle"],
            conf.get("core_product",""), conf.get("recent_news",""),
            conf.get("pain_points",""),  conf.get("pitch_angle",""),
            fi.get("total_raised",""),   fi.get("last_round",""),
            fi.get("key_investors",""),  fi.get("revenue_signals",""), conf.get("funding_info",""),
            "; ".join(js.get("open_roles",[])), js.get("hiring_themes",""),
            js.get("headcount_signal",""), js.get("pitch_implication",""),
            "; ".join(ts.get("tools_identified",[])), ts.get("stack_summary",""),
            ts.get("pitch_implication",""),
            "; ".join(c.get("company_name","") for c in p.get("competitors",[])),
            p.get("search_depth",""), p.get("created_at",""),
        ]
        for ci, v in enumerate(vals, 1):
            ws.cell(row=ri, column=ci, value=v).alignment = Alignment(wrap_text=True, vertical="top")
        for ci, ck in zip([6,7,8,9,14],["core_product","recent_news","pain_points","pitch_angle","funding_info"]):
            ws.cell(row=ri, column=ci).fill = CONF_FILL.get((conf.get(ck) or "low").lower(), CONF_FILL["low"])

    for i, w in enumerate([22,38,38,46,46,10,10,10,10,18,18,26,26,10,
                            30,30,20,30,30,30,30,28,10,16], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
