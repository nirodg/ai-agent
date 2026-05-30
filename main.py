# /// script
# dependencies = [
#   "langchain-core",
#   "langchain-openai",
#   "langchain-ollama",
#   "langgraph",
#   "pydantic",
#   "ddgs",
#   "streamlit",
#   "python-dotenv",
#   "openpyxl",
#   "tenacity",
#   "langsmith",
# ]
# ///

import csv
import io
import json
import os
import shutil
import sqlite3
import time
from datetime import datetime
from pathlib import Path
from typing import Any, List, Optional

from dotenv import load_dotenv

load_dotenv()

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
import streamlit as st
from ddgs import DDGS
from langchain_core.tools import tool
from langchain_openai import ChatOpenAI
from langchain_ollama import ChatOllama
from langgraph.prebuilt import create_react_agent
from langsmith import Client as LangSmithClient
from langsmith import traceable
from pydantic import BaseModel, Field
from tenacity import retry, stop_after_attempt, wait_exponential, retry_if_exception_type
import uuid

# ---------------------------------------------------------
# LANGSMITH SETUP
# LangGraph auto-traces when these env vars are set:
#   LANGSMITH_TRACING=true
#   LANGSMITH_API_KEY=<your key>
#   LANGSMITH_PROJECT=<your project name>
# All three are read from .env via load_dotenv() above.
# We also generate a session ID to group traces per app session.
# ---------------------------------------------------------

LANGSMITH_PROJECT = os.environ.get("LANGSMITH_PROJECT", "sales-enrichment-agent")

if "session_id" not in st.session_state:
    st.session_state["session_id"] = str(uuid.uuid4())

def _ls_metadata(company: str = "", operation: str = "") -> dict:
    """Standard metadata attached to every traceable call."""
    return {
        "session_id":  st.session_state.get("session_id", ""),
        "company":     company,
        "operation":   operation,
        "app_version": "3.0",
    }

# ---------------------------------------------------------
# PAGE CONFIG
# ---------------------------------------------------------

st.set_page_config(
    page_title="AI Sales Enrichment Agent",
    page_icon="🚀",
    layout="wide",
)

# ---------------------------------------------------------
# CONSTANTS & PATHS
# ---------------------------------------------------------

DB_PATH     = Path("enrichment_profiles.db")
BACKUP_DIR  = Path("db_backups")

DEFAULT_MODEL = "openai/gpt-oss-120b:free"
DEFAULT_LLM_PROVIDER = "openrouter"
LLM_PROVIDERS = ("openrouter", "openai", "perplexity", "ollama")
DEFAULT_PROVIDER_MODEL_MAP = {
    "openrouter": "openai/gpt-oss-120b:free",
    "openai": "gpt-5-nano",
    "perplexity": "sonar",
    "ollama": "gpt-oss:120b",
}


def _is_openrouter_free_model(model_name: str) -> bool:
    return (model_name or "").strip().endswith(":free")

DEFAULT_SYSTEM_PROMPT = """\
You are an elite AI Sales Research Agent.

Your responsibilities:
1. ALWAYS use the web search tool.
2. Research the company thoroughly using all provided search queries.
3. Extract:
   - company product / core offering
   - latest news or funding
   - likely pain points
   - sales opportunities
4. Be concise but useful.
5. Never hallucinate fake funding or fake news.
6. Mention uncertainty when data is weak.
"""

PERSONA_PRESETS = {
    "sales_research_strategist": DEFAULT_SYSTEM_PROMPT,
    "pain_point_hunter": """\
You are an elite AI Sales Research Agent focused on uncovering urgent pain points and strong outreach hooks.

Your responsibilities:
1. ALWAYS use the web search tool.
2. Research the company thoroughly using all provided search queries.
3. Extract:
   - company product / core offering
   - latest operational, hiring, or strategic pressure signals
   - likely pain points with emphasis on urgency and business impact
   - crisp sales opportunities tied to those pain points
4. Prioritize specificity, urgency, and concrete hooks for outbound messaging.
5. Never hallucinate fake funding, fake news, or fake operational issues.
6. Mention uncertainty when evidence is weak or indirect.
""",
    "executive_briefing": """\
You are an elite AI Sales Research Agent producing executive-ready briefings for account planning.

Your responsibilities:
1. ALWAYS use the web search tool.
2. Research the company thoroughly using all provided search queries.
3. Extract:
   - company product / core offering
   - latest news or funding with leadership relevance
   - likely strategic risks, growth blockers, and executive priorities
   - sales opportunities framed as concise executive recommendations
4. Write with high signal, sharp prioritization, and leadership-level clarity.
5. Never hallucinate fake funding, fake news, or fake strategic conclusions.
6. Mention uncertainty when data is weak.
""",
}
DEFAULT_PERSONA_PRESET = "sales_research_strategist"
PERSONA_PRESET_LABELS = {
    "sales_research_strategist": "Sales Research Strategist",
    "pain_point_hunter": "Pain Point Hunter",
    "executive_briefing": "Executive Briefing",
}


def get_persona_comparison_note(persona_name: str) -> str:
    persona_name = (persona_name or "").strip().lower()
    notes = {
        "sales_research_strategist": "Balanced baseline analysis with practical sales framing and broad coverage.",
        "pain_point_hunter": "Shifts the angle toward urgent pain points, operational pressure, and sharper outbound hooks.",
        "executive_briefing": "Reframes the same company into a leadership-style summary with strategic priorities and executive implications.",
    }
    return notes.get(persona_name, "Uses a different framing to reinterpret the same analysis from another angle.")

SEARCH_DEPTH_QUERIES = {
    "fast": [
        "{company} company overview",
        "{company} news 2025",
        "{company} funding revenue",
    ],
    "balanced": [
        "{company} company overview product",
        "{company} latest news 2025",
        "{company} funding crunchbase revenue",
        "{company} linkedin employees hiring",
        "{company} competitors market",
    ],
    "deep": [
        "{company} company overview product",
        "{company} latest news press release 2025",
        "{company} funding crunchbase revenue valuation",
        "{company} linkedin employees hiring growth",
        "{company} competitors alternatives",
        "{company} customer reviews pain points problems",
        "{company} technology stack integrations",
        "{company} CEO founder leadership team",
    ],
}

# Step 1 extra query sets (appended on top of depth queries)
JOB_QUERIES      = ["{company} jobs hiring 2025", "{company} open roles engineering sales"]
TECHSTACK_QUERIES= ["{company} tech stack tools integrations", "{company} software uses built with"]

# ---------------------------------------------------------
# DATABASE SETUP
# ---------------------------------------------------------

def get_db() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    with get_db() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS profiles (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                company_name  TEXT NOT NULL,
                core_product  TEXT,
                recent_news   TEXT,
                pain_points   TEXT,
                pitch_angle   TEXT,
                confidence    TEXT,
                funding_info  TEXT,
                competitors   TEXT,
                chat_history  TEXT,
                search_depth  TEXT,
                job_signals   TEXT,
                tech_stack    TEXT,
                intent_score  TEXT,
                created_at    TEXT NOT NULL
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS settings (
                key   TEXT PRIMARY KEY,
                value TEXT
            )
        """)
        conn.execute("""
            INSERT OR IGNORE INTO settings (key, value)
            VALUES ('system_prompt', ?)
        """, (DEFAULT_SYSTEM_PROMPT,))
        conn.execute("""
            INSERT OR IGNORE INTO settings (key, value)
            VALUES ('llm_model_prompt_map', '{}')
        """)
        conn.execute("""
            INSERT OR IGNORE INTO settings (key, value)
            VALUES ('llm_model_persona_preset_map', '{}')
        """)
        conn.execute("""
            INSERT OR IGNORE INTO settings (key, value)
            VALUES ('ai_model', ?)
        """, (DEFAULT_MODEL,))
        conn.execute("""
            INSERT OR IGNORE INTO settings (key, value)
            VALUES ('llm_provider', ?)
        """, (DEFAULT_LLM_PROVIDER,))
        conn.execute("""
            INSERT OR IGNORE INTO settings (key, value)
            VALUES ('llm_provider_model_map', ?)
        """, (json.dumps(DEFAULT_PROVIDER_MODEL_MAP),))
        conn.execute("""
            INSERT OR IGNORE INTO settings (key, value)
            VALUES ('openrouter_api_key', '')
        """)
        conn.execute("""
            INSERT OR IGNORE INTO settings (key, value)
            VALUES ('openrouter_base_url', 'https://openrouter.ai/api/v1')
        """)
        conn.execute("""
            INSERT OR IGNORE INTO settings (key, value)
            VALUES ('pplx_api_key', '')
        """)
        conn.execute("""
            INSERT OR IGNORE INTO settings (key, value)
            VALUES ('pplx_base_url', 'https://api.perplexity.ai')
        """)
        conn.execute("""
            INSERT OR IGNORE INTO settings (key, value)
            VALUES ('ollama_api_key', '')
        """)
        conn.execute("""
            INSERT OR IGNORE INTO settings (key, value)
            VALUES ('ollama_base_url', 'https://ollama.com')
        """)
        conn.execute("""
            INSERT OR IGNORE INTO settings (key, value)
            VALUES ('ollama_insecure_ssl', 'true')
        """)
        conn.execute("""
            INSERT OR IGNORE INTO settings (key, value)
            VALUES ('llm_timeout', '30')
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS email_drafts (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                company_name  TEXT NOT NULL,
                subject       TEXT,
                body          TEXT NOT NULL,
                tone          TEXT,
                is_enhanced   INTEGER DEFAULT 0,
                base_draft_id INTEGER,
                created_at    TEXT NOT NULL
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS notes (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                company_name  TEXT NOT NULL,
                note          TEXT NOT NULL,
                created_at    TEXT NOT NULL
            )
        """)
        # Migrations — safe to run on existing DBs
        for col in ("confidence", "funding_info", "competitors", "search_depth",
                    "job_signals", "tech_stack", "intent_score"):
            try:
                conn.execute(f"ALTER TABLE profiles ADD COLUMN {col} TEXT")
            except Exception:
                pass
        conn.commit()


init_db()

# ---------------------------------------------------------
# DB HELPERS — PROFILES
# ---------------------------------------------------------

def save_profile(profile_dict: dict, confidence: dict, funding_info: dict,
                 competitors: list, chat_history: list, search_depth: str,
                 job_signals: dict = None, tech_stack: dict = None) -> int:
    with get_db() as conn:
        cur = conn.execute("""
            INSERT INTO profiles
                (company_name, core_product, recent_news, pain_points, pitch_angle,
                 confidence, funding_info, competitors, chat_history, search_depth,
                 job_signals, tech_stack, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            profile_dict["company_name"],
            profile_dict["core_product"],
            profile_dict["recent_news"],
            json.dumps(profile_dict["pain_points"]),
            profile_dict["pitch_angle"],
            json.dumps(confidence),
            json.dumps(funding_info),
            json.dumps(competitors),
            json.dumps(chat_history),
            search_depth,
            json.dumps(job_signals or {}),
            json.dumps(tech_stack  or {}),
            datetime.now().strftime("%Y-%m-%d %H:%M"),
        ))
        conn.commit()
        return cur.lastrowid


def load_all_profiles() -> list:
    with get_db() as conn:
        return conn.execute(
            "SELECT * FROM profiles ORDER BY created_at DESC"
        ).fetchall()


def delete_profile(profile_id: int):
    with get_db() as conn:
        conn.execute("DELETE FROM profiles WHERE id = ?", (profile_id,))
        conn.commit()


def row_to_dict(row) -> dict:
    d = dict(row)
    d["pain_points"]  = json.loads(d["pain_points"]   or "[]")
    d["chat_history"] = json.loads(d["chat_history"]  or "[]")
    d["confidence"]   = json.loads(d["confidence"]    or "{}")
    d["funding_info"] = json.loads(d["funding_info"]  or "{}")
    d["competitors"]  = json.loads(d["competitors"]   or "[]")
    d["job_signals"]  = json.loads(d["job_signals"]   or "{}")
    d["tech_stack"]   = json.loads(d["tech_stack"]    or "{}")
    d["intent_score"] = json.loads(d["intent_score"]  or "null") if d.get("intent_score") else None
    return d


def save_intent_score(profile_id: int, score_dict: dict):
    with get_db() as conn:
        conn.execute(
            "UPDATE profiles SET intent_score = ? WHERE id = ?",
            (json.dumps(score_dict), profile_id),
        )
        conn.commit()


# ---------------------------------------------------------
# DB HELPERS — SETTINGS
# ---------------------------------------------------------

def load_setting(key: str) -> str:
    with get_db() as conn:
        row = conn.execute("SELECT value FROM settings WHERE key = ?", (key,)).fetchone()
        return row["value"] if row else ""


def save_setting(key: str, value: str):
    with get_db() as conn:
        conn.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)", (key, value))
        conn.commit()


# ---------------------------------------------------------
# LLM SETTINGS & FACTORY
# ---------------------------------------------------------

def _to_int(raw: str, fallback: int) -> int:
    try:
        return int(str(raw).strip())
    except Exception:
        return fallback


def get_llm_provider() -> str:
    raw = (load_setting("llm_provider") or os.getenv("LLM_PROVIDER") or DEFAULT_LLM_PROVIDER).strip().lower()
    return raw if raw in LLM_PROVIDERS else DEFAULT_LLM_PROVIDER


def get_provider_model_map() -> dict:
    raw = load_setting("llm_provider_model_map")
    model_map = {}
    if raw:
        try:
            parsed = json.loads(raw)
            if isinstance(parsed, dict):
                model_map = {str(k): str(v).strip() for k, v in parsed.items() if str(v).strip()}
        except Exception:
            model_map = {}

    changed = False
    for provider, default_model in DEFAULT_PROVIDER_MODEL_MAP.items():
        if not model_map.get(provider):
            model_map[provider] = default_model
            changed = True

    if model_map.get("openrouter") and not _is_openrouter_free_model(model_map["openrouter"]):
        model_map["openrouter"] = DEFAULT_PROVIDER_MODEL_MAP["openrouter"]
        changed = True

    # Silent migration path from legacy single-model setting.
    legacy_model = (load_setting("ai_model") or os.getenv("LLM_MODEL") or "").strip()
    current_provider = get_llm_provider()
    if legacy_model and current_provider != "openrouter" and model_map.get(current_provider) != legacy_model:
        model_map[current_provider] = legacy_model
        changed = True

    if changed:
        save_setting("llm_provider_model_map", json.dumps(model_map))

    return model_map


def save_llm_model_for_provider(provider: str, model_name: str):
    provider = (provider or "").strip().lower()
    model_name = (model_name or "").strip()
    if provider not in LLM_PROVIDERS or not model_name:
        return
    if provider == "openrouter" and not _is_openrouter_free_model(model_name):
        return
    model_map = get_provider_model_map()
    model_map[provider] = model_name
    save_setting("llm_provider_model_map", json.dumps(model_map))
    if provider == get_llm_provider():
        save_setting("ai_model", model_name)


def get_llm_model(provider: Optional[str] = None) -> str:
    provider = (provider or get_llm_provider()).strip().lower()
    model_map = get_provider_model_map()
    scoped = (model_map.get(provider) or "").strip()
    if provider == "openrouter" and scoped and not _is_openrouter_free_model(scoped):
        scoped = ""
    if scoped:
        return scoped
    provider_default = (DEFAULT_PROVIDER_MODEL_MAP.get(provider) or "").strip()
    if provider_default:
        return provider_default
    return (load_setting("ai_model") or os.getenv("LLM_MODEL") or DEFAULT_MODEL).strip() or DEFAULT_MODEL


def get_openrouter_api_key() -> str:
    return (
        load_setting("openrouter_api_key")
        or os.getenv("OPENROUTER_API_KEY")
        or ""
    ).strip()


def get_openrouter_base_url() -> str:
    return (
        load_setting("openrouter_base_url")
        or os.getenv("OPENROUTER_BASE_URL")
        or "https://openrouter.ai/api/v1"
    ).strip()


def _model_prompt_key(provider: Optional[str] = None, model_name: Optional[str] = None) -> str:
    resolved_provider = (provider or get_llm_provider()).strip().lower()
    resolved_model = (model_name or get_llm_model(resolved_provider)).strip()
    return f"{resolved_provider}:{resolved_model}"


def _load_json_setting_map(key: str) -> dict[str, str]:
    raw = load_setting(key)
    if not raw:
        return {}
    try:
        parsed = json.loads(raw)
    except Exception:
        return {}
    if not isinstance(parsed, dict):
        return {}
    return {str(map_key): str(map_value) for map_key, map_value in parsed.items() if str(map_value).strip()}


def get_model_persona_preset_map() -> dict[str, str]:
    preset_map = _load_json_setting_map("llm_model_persona_preset_map")
    changed = False
    for key, preset_name in list(preset_map.items()):
        normalized = preset_name.strip().lower()
        if normalized not in PERSONA_PRESETS:
            preset_map[key] = DEFAULT_PERSONA_PRESET
            changed = True
    if changed:
        save_setting("llm_model_persona_preset_map", json.dumps(preset_map))
    return preset_map


def save_persona_preset_for_model(provider: str, model_name: str, preset_name: str):
    normalized_preset = (preset_name or "").strip().lower()
    if normalized_preset not in PERSONA_PRESETS:
        return
    preset_map = get_model_persona_preset_map()
    preset_map[_model_prompt_key(provider, model_name)] = normalized_preset
    save_setting("llm_model_persona_preset_map", json.dumps(preset_map))


def get_persona_preset_for_model(provider: Optional[str] = None, model_name: Optional[str] = None) -> str:
    key = _model_prompt_key(provider, model_name)
    preset_name = get_model_persona_preset_map().get(key, DEFAULT_PERSONA_PRESET).strip().lower()
    return preset_name if preset_name in PERSONA_PRESETS else DEFAULT_PERSONA_PRESET


def get_model_prompt_map() -> dict[str, str]:
    prompt_map = _load_json_setting_map("llm_model_prompt_map")
    legacy_prompt = (load_setting("system_prompt") or "").strip()
    current_key = _model_prompt_key()
    if legacy_prompt and legacy_prompt != DEFAULT_SYSTEM_PROMPT and not prompt_map.get(current_key):
        prompt_map[current_key] = legacy_prompt
        save_setting("llm_model_prompt_map", json.dumps(prompt_map))
    return prompt_map


def save_system_prompt_for_model(provider: str, model_name: str, prompt: str):
    cleaned_prompt = (prompt or "").strip()
    key = _model_prompt_key(provider, model_name)
    prompt_map = get_model_prompt_map()
    if cleaned_prompt:
        prompt_map[key] = cleaned_prompt
    else:
        prompt_map.pop(key, None)
    save_setting("llm_model_prompt_map", json.dumps(prompt_map))


def reset_system_prompt_for_model(provider: str, model_name: str):
    key = _model_prompt_key(provider, model_name)
    prompt_map = get_model_prompt_map()
    prompt_map.pop(key, None)
    save_setting("llm_model_prompt_map", json.dumps(prompt_map))


def get_current_system_prompt(provider: Optional[str] = None, model_name: Optional[str] = None) -> str:
    key = _model_prompt_key(provider, model_name)
    prompt_map = get_model_prompt_map()
    custom_prompt = (prompt_map.get(key) or "").strip()
    if custom_prompt:
        return custom_prompt
    preset_name = get_persona_preset_for_model(provider, model_name)
    return PERSONA_PRESETS.get(preset_name, DEFAULT_SYSTEM_PROMPT)


def invoke_with_provider_fallback(prompt: str, temperature: float, where: str) -> str:
    model = get_llm_client(temperature=temperature)
    return model.invoke(prompt).content.strip()


def get_llm_timeout() -> int:
    raw = (load_setting("llm_timeout") or os.getenv("LLM_TIMEOUT") or "30").strip()
    return _to_int(raw, 30)


def get_perplexity_api_key() -> str:
    return (
        load_setting("pplx_api_key")
        or os.getenv("PPLX_API_KEY")
        or os.getenv("PERPLEXITY_API_KEY")
        or ""
    ).strip()


def get_perplexity_base_url() -> str:
    return (load_setting("pplx_base_url") or os.getenv("PPLX_BASE_URL") or "https://api.perplexity.ai").strip()


def get_ollama_api_key() -> str:
    return (
        os.getenv("OLLAMA_CLOUD_API_KEY")
        or os.getenv("OLLAMA_API_KEY")
        or load_setting("ollama_api_key")
        or ""
    ).strip()
    
def get_ollama_api_key_source() -> str:
    if os.getenv("OLLAMA_CLOUD_API_KEY"):
        return "environment: OLLAMA_CLOUD_API_KEY"
    if os.getenv("OLLAMA_API_KEY"):
        return "environment: OLLAMA_API_KEY"
    if load_setting("ollama_api_key"):
        return "saved setting: ollama_api_key"
    return "missing"


def get_ollama_base_url() -> str:
    return (
        load_setting("ollama_base_url")
        or os.getenv("OLLAMA_BASE_URL")
        or os.getenv("OLLAMA_CLOUD_BASE_URL")
        or "https://ollama.com"
    ).strip()


def get_ollama_insecure_ssl() -> bool:
    raw_value = (
        load_setting("ollama_insecure_ssl")
        or os.getenv("OLLAMA_INSECURE_SSL")
        or "true"
    ).strip().lower()
    return raw_value in {"1", "true", "yes", "on"}


def get_ollama_client_kwargs(timeout_seconds: int) -> dict[str, Any]:
    client_kwargs: dict[str, Any] = {"timeout": timeout_seconds}
    api_key = get_ollama_api_key()
    if api_key:
        client_kwargs["headers"] = {"Authorization": f"Bearer {api_key}"}
    if get_ollama_insecure_ssl():
        client_kwargs["verify"] = False
    return client_kwargs


def test_ollama_connection() -> str:
    model = ChatOllama(
        model=get_llm_model("ollama"),
        base_url=get_ollama_base_url(),
        temperature=0.0,
        client_kwargs=get_ollama_client_kwargs(get_llm_timeout()),
    )
    response = model.invoke("Reply with exactly: OK")
    return response.content.strip()


def get_llm_client(temperature: float = 0.0):
    provider = get_llm_provider()
    model_name = get_llm_model(provider)
    timeout_seconds = get_llm_timeout()

    if provider == "openrouter":
        return ChatOpenAI(
            model=model_name,
            temperature=temperature,
            base_url=get_openrouter_base_url(),
            api_key=get_openrouter_api_key() or "missing-openrouter-key",
            request_timeout=timeout_seconds,
        )

    if provider == "perplexity":
        return ChatOpenAI(
            model=model_name,
            temperature=temperature,
            base_url=get_perplexity_base_url(),
            api_key=get_perplexity_api_key() or "missing-pplx-key",
            request_timeout=timeout_seconds,
        )

    if provider == "ollama":
        return ChatOllama(
            model=model_name,
            temperature=temperature,
            base_url=get_ollama_base_url(),
            client_kwargs=get_ollama_client_kwargs(timeout_seconds),
        )

    return ChatOpenAI(
        model=model_name,
        temperature=temperature,
        request_timeout=timeout_seconds,
    )


# ---------------------------------------------------------
# DB HELPERS — EMAIL DRAFTS
# ---------------------------------------------------------

def save_email_draft(company_name: str, subject: str, body: str,
                     tone: str, is_enhanced: bool = False,
                     base_draft_id: Optional[int] = None) -> int:
    with get_db() as conn:
        cur = conn.execute("""
            INSERT INTO email_drafts
                (company_name, subject, body, tone, is_enhanced, base_draft_id, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (
            company_name, subject, body, tone,
            1 if is_enhanced else 0,
            base_draft_id,
            datetime.now().strftime("%Y-%m-%d %H:%M"),
        ))
        conn.commit()
        return cur.lastrowid


def load_drafts_for_company(company_name: str) -> list:
    with get_db() as conn:
        return conn.execute("""
            SELECT * FROM email_drafts WHERE company_name = ?
            ORDER BY created_at DESC
        """, (company_name,)).fetchall()


def delete_email_draft(draft_id: int):
    with get_db() as conn:
        conn.execute("DELETE FROM email_drafts WHERE id = ?", (draft_id,))
        conn.commit()


# ---------------------------------------------------------
# DB HELPERS — NOTES  (Step 3)
# ---------------------------------------------------------

def save_note(company_name: str, note: str) -> int:
    with get_db() as conn:
        cur = conn.execute("""
            INSERT INTO notes (company_name, note, created_at)
            VALUES (?, ?, ?)
        """, (company_name, note, datetime.now().strftime("%Y-%m-%d %H:%M")))
        conn.commit()
        return cur.lastrowid


def load_notes(company_name: str) -> list:
    with get_db() as conn:
        return conn.execute("""
            SELECT * FROM notes WHERE company_name = ?
            ORDER BY created_at DESC
        """, (company_name,)).fetchall()


def delete_note(note_id: int):
    with get_db() as conn:
        conn.execute("DELETE FROM notes WHERE id = ?", (note_id,))
        conn.commit()


# ---------------------------------------------------------
# DB BACKUP & RESTORE  (Step 3)
# ---------------------------------------------------------

def create_backup() -> Path:
    BACKUP_DIR.mkdir(exist_ok=True)
    ts      = datetime.now().strftime("%Y%m%d_%H%M%S")
    dst     = BACKUP_DIR / f"enrichment_profiles_{ts}.db"
    shutil.copy2(DB_PATH, dst)
    return dst


def restore_backup(uploaded_bytes: bytes):
    DB_PATH.write_bytes(uploaded_bytes)


def list_backups() -> list[Path]:
    if not BACKUP_DIR.exists():
        return []
    return sorted(BACKUP_DIR.glob("*.db"), reverse=True)


# ---------------------------------------------------------
# MEMORY BLOCK
# ---------------------------------------------------------

def build_company_memory_block(profiles: list) -> str:
    if not profiles:
        return ""
    lines = [
        "\n\n---\n## 🧠 Companies You Have Already Researched\n",
        "Use this knowledge to avoid repetition and draw comparisons:\n",
    ]
    for row in profiles:
        p = row_to_dict(row)
        pain_summary = "; ".join(p["pain_points"][:2])
        lines.append(
            f"- **{p['company_name']}** ({p['created_at']}): "
            f"{p['core_product']}. Key pains: {pain_summary}."
        )
    return "\n".join(lines)


# ---------------------------------------------------------
# STRUCTURED OUTPUT SCHEMAS
# ---------------------------------------------------------

class FieldConfidence(BaseModel):
    core_product: str = Field(description="Confidence level: high, medium, or low")
    recent_news:  str = Field(description="Confidence level: high, medium, or low")
    pain_points:  str = Field(description="Confidence level: high, medium, or low")
    pitch_angle:  str = Field(description="Confidence level: high, medium, or low")
    funding_info: str = Field(description="Confidence level: high, medium, or low")


class FundingInfo(BaseModel):
    total_raised:    str = Field(description="Total funding raised. Use 'Unknown' if not found.")
    last_round:      str = Field(description="Most recent round type and date. Use 'Unknown' if not found.")
    key_investors:   str = Field(description="Notable investors, comma-separated. Use 'Unknown' if not found.")
    revenue_signals: str = Field(description="Revenue, ARR, or headcount signals. Use 'Unknown' if not found.")


class CompetitorProfile(BaseModel):
    company_name: str            = Field(description="Competitor company name.")
    core_product: str            = Field(description="What the competitor sells or does.")
    recent_news:  str            = Field(description="Key recent update for the competitor.")
    pain_points:  List[str]      = Field(description="Business challenges this competitor faces.")
    pitch_angle:  str            = Field(description="How to differentiate against this competitor.")
    confidence:   FieldConfidence = Field(description="Confidence per field.")


class JobSignals(BaseModel):
    open_roles:        List[str] = Field(description="List of notable open roles found (e.g. 'Senior Data Engineer', 'VP Sales'). Empty list if none found.")
    hiring_themes:     str       = Field(description="What the hiring patterns suggest about company priorities. Use 'Unknown' if not found.")
    headcount_signal:  str       = Field(description="Any headcount size or growth signals from job postings. Use 'Unknown' if not found.")
    pitch_implication: str       = Field(description="How these hiring signals should shape the sales pitch.")


class TechStack(BaseModel):
    tools_identified: List[str] = Field(description="List of tools/platforms/frameworks identified (e.g. 'Salesforce', 'AWS', 'React'). Empty list if none found.")
    stack_summary:    str       = Field(description="Brief summary of what the stack reveals about the company's technical maturity. Use 'Unknown' if not found.")
    pitch_implication:str       = Field(description="How the tech stack should shape the sales pitch.")


class CompanyProfile(BaseModel):
    company_name: str                     = Field(description="The formal name of the company.")
    core_product: str                     = Field(description="Summary of what the company sells or does.")
    recent_news:  str                     = Field(description="Key recent update, funding, or press release.")
    pain_points:  List[str]               = Field(description="Business challenges this company likely faces.")
    pitch_angle:  str                     = Field(description="Tailored value proposition for outreach.")
    confidence:   FieldConfidence         = Field(description="Confidence per field.")
    funding_info: FundingInfo             = Field(description="Funding and financial signals.")
    competitors:  List[CompetitorProfile] = Field(description="2-3 direct competitors with full profiles.")
    job_signals:  JobSignals              = Field(description="Job posting analysis and hiring signals.")
    tech_stack:   TechStack               = Field(description="Inferred technology stack and tools.")


# ---------------------------------------------------------
# RATE-LIMITED SEARCH TOOL  (Step 3 — retry + backoff)
# ---------------------------------------------------------

@retry(
    retry=retry_if_exception_type(Exception),
    wait=wait_exponential(multiplier=1, min=2, max=30),
    stop=stop_after_attempt(3),
    reraise=True,
)
def _ddgs_search(query: str, max_results: int = 5) -> list:
    with DDGS() as client:
        return list(client.text(query, max_results=max_results))


@tool
def free_duckduckgo_search(query: str) -> str:
    """Search the web for company information and news."""
    try:
        results = _ddgs_search(query)
        if not results:
            return json.dumps({"query": query, "results": []})
        cleaned = [
            {"title": r.get("title", ""), "snippet": r.get("body", ""), "url": r.get("href", "")}
            for r in results
        ]
        return json.dumps(cleaned, indent=2)
    except Exception as e:
        return json.dumps({"error": str(e), "query": query})


# ---------------------------------------------------------
# AGENT FACTORY
# ---------------------------------------------------------

def build_agent(system_prompt: str, temperature: float = 0.0):
    model = get_llm_client(temperature=temperature)
    agent = create_react_agent(
        model=model,
        tools=[free_duckduckgo_search],
        prompt=system_prompt,
    )
    return agent, model


# ---------------------------------------------------------
# STEP 1 — JOB & TECH STACK RESEARCH
# ---------------------------------------------------------

@traceable(name="research_job_signals", tags=["step1", "jobs"])
def research_job_signals(company_name: str) -> dict:
    """Run dedicated job-posting searches and extract structured signals."""
    search_results = []
    for q_tmpl in JOB_QUERIES:
        try:
            results = _ddgs_search(q_tmpl.format(company=company_name), max_results=5)
            search_results.extend(results)
        except Exception:
            pass

    if not search_results:
        return {
            "open_roles": [], "hiring_themes": "Unknown",
            "headcount_signal": "Unknown", "pitch_implication": "Unknown",
        }

    snippets = "\n".join(
        f"- {r.get('title', '')}: {r.get('body', '')[:200]}"
        for r in search_results[:8]
    )
    prompt = f"""
Analyse these job posting search results for {company_name} and extract structured hiring signals.
Respond ONLY with a JSON object matching this schema — no markdown, no preamble:
{{
  "open_roles": ["role1", "role2"],
  "hiring_themes": "...",
  "headcount_signal": "...",
  "pitch_implication": "..."
}}

SEARCH RESULTS:
{snippets}
"""
    try:
        raw = invoke_with_provider_fallback(prompt, temperature=0, where="research_job_signals").replace("```json","").replace("```","")
        return json.loads(raw)
    except Exception:
        return {
            "open_roles": [], "hiring_themes": "Unknown",
            "headcount_signal": "Unknown", "pitch_implication": "Unknown",
        }


@traceable(name="research_tech_stack", tags=["step1", "tech-stack"])
def research_tech_stack(company_name: str) -> dict:
    """Run dedicated tech-stack searches and extract structured signals."""
    search_results = []
    for q_tmpl in TECHSTACK_QUERIES:
        try:
            results = _ddgs_search(q_tmpl.format(company=company_name), max_results=5)
            search_results.extend(results)
        except Exception:
            pass

    if not search_results:
        return {
            "tools_identified": [], "stack_summary": "Unknown",
            "pitch_implication": "Unknown",
        }

    snippets = "\n".join(
        f"- {r.get('title', '')}: {r.get('body', '')[:200]}"
        for r in search_results[:8]
    )
    prompt = f"""
Analyse these search results for {company_name} and extract their technology stack and tools.
Respond ONLY with a JSON object — no markdown, no preamble:
{{
  "tools_identified": ["tool1", "tool2"],
  "stack_summary": "...",
  "pitch_implication": "..."
}}

SEARCH RESULTS:
{snippets}
"""
    try:
        raw = invoke_with_provider_fallback(prompt, temperature=0, where="research_tech_stack").replace("```json","").replace("```","")
        return json.loads(raw)
    except Exception:
        return {
            "tools_identified": [], "stack_summary": "Unknown",
            "pitch_implication": "Unknown",
        }


# ---------------------------------------------------------
# STEP 1 — TRIGGER ALERT / RE-RESEARCH
# ---------------------------------------------------------

def diff_profiles(old: dict, new: dict) -> list[str]:
    """Return a list of human-readable change descriptions between two profiles."""
    changes = []

    def changed(key):
        return str(old.get(key, "")) != str(new.get(key, ""))

    if changed("recent_news"):
        changes.append(f"📰 **News updated:** {new.get('recent_news', '')[:200]}")

    old_fi = old.get("funding_info", {})
    new_fi = new.get("funding_info", {})
    for fk, label in [("total_raised","💰 Funding"), ("last_round","📅 Last round"), ("key_investors","🏦 Investors")]:
        if str(old_fi.get(fk,"")) != str(new_fi.get(fk,"")):
            changes.append(f"{label} changed: {new_fi.get(fk,'Unknown')}")

    old_pains = set(old.get("pain_points", []))
    new_pains = set(new.get("pain_points", []))
    added = new_pains - old_pains
    if added:
        changes.append(f"🎯 New pain points identified: {'; '.join(added)}")

    old_jobs = old.get("job_signals", {})
    new_jobs = new.get("job_signals", {})
    if str(old_jobs.get("hiring_themes","")) != str(new_jobs.get("hiring_themes","")):
        changes.append(f"👥 Hiring themes changed: {new_jobs.get('hiring_themes','Unknown')}")

    old_stack = set(old.get("tech_stack", {}).get("tools_identified", []))
    new_stack = set(new.get("tech_stack", {}).get("tools_identified", []))
    added_tools = new_stack - old_stack
    if added_tools:
        changes.append(f"🛠 New tools detected: {', '.join(added_tools)}")

    return changes if changes else ["No significant changes detected."]


# ---------------------------------------------------------
# ENRICHMENT CORE
# ---------------------------------------------------------

def build_enrichment_prompt(company_name: str, depth: str) -> str:
    queries   = SEARCH_DEPTH_QUERIES[depth]
    formatted = "\n".join(f"{i+1}. {q.format(company=company_name)}" for i, q in enumerate(queries))
    return f"""Research the company: {company_name}

You MUST run ALL of the following search queries before answering.
Do not skip any. Each covers a different data source angle:

{formatted}

After completing all searches, synthesize the results into a comprehensive report covering:
- Core product and business model
- Recent news, press releases, or announcements
- Funding rounds, revenue estimates, headcount signals
- Key pain points and business challenges
- 2-3 direct competitors (research each one too)
- Sales pitch angle

Be specific. Cite uncertainty where data is missing or weak.
"""


@traceable(name="enrich_company", tags=["enrichment", "core"])
@retry(
    retry=retry_if_exception_type(Exception),
    wait=wait_exponential(multiplier=1, min=2, max=60),
    stop=stop_after_attempt(3),
    reraise=True,
)
def enrich_company(company_name: str, full_system_prompt: str, depth: str,
                   include_jobs: bool = True, include_stack: bool = True):
    """Full enrichment pipeline with optional job + stack research."""
    agent, base_model = build_agent(full_system_prompt)
    user_message      = build_enrichment_prompt(company_name, depth)
    result = agent.invoke({"messages": [{"role": "user", "content": user_message}]})
    raw_response      = result["messages"][-1].content

    structured_model  = base_model.with_structured_output(CompanyProfile)
    final_profile: CompanyProfile = structured_model.invoke(
        f"Extract a fully structured company profile from the following research report.\n\nREPORT:\n{raw_response}"
    )

    d            = final_profile.model_dump()
    confidence   = d.pop("confidence")
    funding_info = d.pop("funding_info")
    competitors  = d.pop("competitors")
    job_signals  = d.pop("job_signals")
    tech_stack   = d.pop("tech_stack")

    for comp in competitors:
        comp["confidence"] = comp.pop("confidence", {})

    # Step 1 — dedicated deep searches (override LLM-inferred values)
    if include_jobs:
        job_signals = research_job_signals(company_name)
    if include_stack:
        tech_stack  = research_tech_stack(company_name)

    return d, confidence, funding_info, competitors, job_signals, tech_stack


# ---------------------------------------------------------
# INTENT SCORING ENGINE
# ---------------------------------------------------------

INTENT_SIGNALS = [
    ("Recent funding round mentioned",
     lambda p: any(w in (p.get("recent_news") or "").lower()
                   for w in ["series","seed","raised","funding","investment","round"]), 25),
    ("Funding total raised is known",
     lambda p: (p.get("funding_info") or {}).get("total_raised","Unknown").lower() != "unknown", 10),
    ("Revenue / headcount signals found",
     lambda p: (p.get("funding_info") or {}).get("revenue_signals","Unknown").lower() != "unknown", 10),
    ("3+ pain points identified",
     lambda p: len(p.get("pain_points") or []) >= 3, 15),
    ("High-confidence core product data",
     lambda p: (p.get("confidence") or {}).get("core_product","low") == "high", 10),
    ("High-confidence recent news",
     lambda p: (p.get("confidence") or {}).get("recent_news","low") == "high", 10),
    ("Researched at deep depth",
     lambda p: p.get("search_depth") == "deep", 10),
    ("Competitors mapped",
     lambda p: len(p.get("competitors") or []) >= 2, 10),
    ("Hiring signals found",
     lambda p: len((p.get("job_signals") or {}).get("open_roles", [])) > 0, 5),
    ("Tech stack identified",
     lambda p: len((p.get("tech_stack") or {}).get("tools_identified", [])) > 0, 5),
]


def compute_rule_score(profile: dict) -> tuple[int, list[str]]:
    score, signals = 0, []
    for desc, test, pts in INTENT_SIGNALS:
        try:
            if test(profile):
                score += pts
                signals.append(desc)
        except Exception:
            pass
    return min(score, 100), signals


@traceable(name="compute_intent_score", tags=["scoring", "intent"])
def compute_intent_score(profile: dict) -> dict:
    rule_score, signals = compute_rule_score(profile)
    prompt = f"""
You are a B2B sales intelligence analyst.
Rule-based score: {rule_score}/100. Triggered signals:
{chr(10).join(f"- {s}" for s in signals) if signals else "- None"}

PROFILE:
- Company: {profile.get('company_name')}
- Product: {profile.get('core_product')}
- News: {profile.get('recent_news')}
- Pains: {"; ".join(profile.get('pain_points') or [])}
- Hiring: {json.dumps(profile.get('job_signals') or {})}
- Stack: {json.dumps(profile.get('tech_stack') or {})}
- Funding: {json.dumps(profile.get('funding_info') or {})}

Respond ONLY with JSON (no markdown):
{{
  "adjusted_score": <1-10>,
  "reasoning": "<2-3 sentences>",
  "recommended_action": "<one concrete next step>",
  "best_time_to_reach": "<timing guidance>"
}}
"""
    try:
        raw    = invoke_with_provider_fallback(prompt, temperature=0, where="compute_intent_score").replace("```json","").replace("```","")
        llm_out = json.loads(raw)
    except Exception:
        llm_out = {
            "adjusted_score": max(1, rule_score // 10),
            "reasoning": "LLM scoring unavailable.",
            "recommended_action": "Review profile manually.",
            "best_time_to_reach": "Unknown",
        }
    return {
        "rule_score":         rule_score,
        "score":              int(llm_out.get("adjusted_score", max(1, rule_score // 10))),
        "signals":            signals,
        "reasoning":          llm_out.get("reasoning", ""),
        "recommended_action": llm_out.get("recommended_action", ""),
        "best_time_to_reach": llm_out.get("best_time_to_reach", ""),
    }


# ---------------------------------------------------------
# STEP 2 — OUTREACH GENERATION
# ---------------------------------------------------------

TONE_INSTRUCTIONS = {
    "formal":   "Write in a formal, executive business tone. Full sentences, no slang. Suitable for C-suite.",
    "friendly": "Write in a warm, conversational tone. Approachable and human. No jargon.",
    "bold":     "Write in a direct, punchy tone. Short sentences. Strong value claim upfront. No fluff.",
}


def _base_email_context(profile: dict) -> str:
    pains = "\n".join(f"- {p}" for p in profile.get("pain_points", []))
    jobs  = profile.get("job_signals", {})
    stack = profile.get("tech_stack", {})
    extra = ""
    if jobs.get("pitch_implication") and jobs["pitch_implication"] != "Unknown":
        extra += f"\n- Hiring insight: {jobs['pitch_implication']}"
    if stack.get("pitch_implication") and stack["pitch_implication"] != "Unknown":
        extra += f"\n- Tech stack insight: {stack['pitch_implication']}"
    return f"""
Company: {profile.get('company_name')}
Core product: {profile.get('core_product')}
Recent news: {profile.get('recent_news')}
Pain points:
{pains}
Pitch angle: {profile.get('pitch_angle')}{extra}
""".strip()


@traceable(name="generate_email_draft", tags=["step2", "outreach", "email"])
def generate_email_draft(profile: dict, tone: str) -> str:
    prompt = f"""
You are an expert B2B sales copywriter.
Write a cold outreach email to a decision-maker at {profile['company_name']}.

{_base_email_context(profile)}

TONE: {TONE_INSTRUCTIONS.get(tone, TONE_INSTRUCTIONS['formal'])}
RULES:
- Subject line first, prefixed "Subject: "
- Blank line then email body
- Under 200 words
- Single CTA (e.g. 15-min call)
- Sign off as "The Team"
- No "I hope this finds you well"
"""
    return invoke_with_provider_fallback(prompt, temperature=0.7, where="generate_email_draft")


@traceable(name="generate_followup_sequence", tags=["step2", "outreach", "sequence"])
def generate_followup_sequence(profile: dict, tone: str) -> list[dict]:
    """Generate a 3-touch follow-up sequence."""
    prompt = f"""
You are an expert B2B sales copywriter.
Write a 3-email follow-up sequence for {profile['company_name']}.

{_base_email_context(profile)}

TONE: {TONE_INSTRUCTIONS.get(tone, TONE_INSTRUCTIONS['formal'])}

Return ONLY a JSON array with exactly 3 objects — no markdown:
[
  {{"touch": 1, "send_gap": "Day 1", "subject": "...", "body": "..."}},
  {{"touch": 2, "send_gap": "Day 4", "subject": "...", "body": "..."}},
  {{"touch": 3, "send_gap": "Day 10", "subject": "...", "body": "..."}}
]
Each email under 150 words. Touch 3 is a short "break-up" email.
Sign off as "The Team". No placeholders.
"""
    try:
        raw = invoke_with_provider_fallback(prompt, temperature=0.7, where="generate_followup_sequence").replace("```json","").replace("```","")
        return json.loads(raw)
    except Exception:
        return []


@traceable(name="generate_linkedin_message", tags=["step2", "outreach", "linkedin"])
def generate_linkedin_message(profile: dict) -> str:
    """Generate a short LinkedIn DM variant."""
    prompt = f"""
Write a LinkedIn cold outreach message for {profile['company_name']}.

{_base_email_context(profile)}

RULES:
- Max 300 characters (LinkedIn connection note limit)
- Conversational, no corporate jargon
- One clear hook + one soft CTA
- No subject line needed
- No sign-off name
"""
    return invoke_with_provider_fallback(prompt, temperature=0.7, where="generate_linkedin_message")


@traceable(name="generate_objection_prep", tags=["step2", "outreach", "objections"])
def generate_objection_prep(profile: dict) -> list[dict]:
    """Generate top 3 objections + rebuttals."""
    prompt = f"""
You are a B2B sales coach.
For {profile['company_name']}, generate the 3 most likely objections
a prospect would raise and a sharp rebuttal for each.

{_base_email_context(profile)}

Return ONLY a JSON array — no markdown:
[
  {{"objection": "...", "rebuttal": "..."}},
  {{"objection": "...", "rebuttal": "..."}},
  {{"objection": "...", "rebuttal": "..."}}
]
"""
    try:
        raw = invoke_with_provider_fallback(prompt, temperature=0.3, where="generate_objection_prep").replace("```json","").replace("```","")
        return json.loads(raw)
    except Exception:
        return []


@traceable(name="generate_enhanced_draft", tags=["step2", "outreach", "enhance"])
def generate_enhanced_draft(profile: dict, base_subject: str,
                             base_body: str, instructions: str, tone: str) -> str:
    prompt = f"""
Improve this cold email draft for {profile['company_name']}.

ORIGINAL:
Subject: {base_subject}
{base_body}

ENHANCEMENT INSTRUCTIONS: {instructions}
TONE: {TONE_INSTRUCTIONS.get(tone, TONE_INSTRUCTIONS['formal'])}

Output: Subject line (prefixed "Subject: "), blank line, improved body. Under 200 words. Sign off "The Team".
"""
    return invoke_with_provider_fallback(prompt, temperature=0.7, where="generate_enhanced_draft")


def parse_email(raw: str) -> tuple[str, str]:
    lines, subject, body_lines = raw.splitlines(), "", []
    for line in lines:
        if line.lower().startswith("subject:"):
            subject = line.split(":", 1)[-1].strip()
        else:
            body_lines.append(line)
    return subject, "\n".join(body_lines).strip()


def render_diff(original: str, enhanced: str):
    import difflib
    differ   = difflib.HtmlDiff(wrapcolumn=60)
    diff_html = differ.make_table(
        original.splitlines(), enhanced.splitlines(),
        fromdesc="Original", todesc="Enhanced", context=True, numlines=2,
    )
    st.markdown(f"""
    <style>
      .diff{{font-family:monospace;font-size:.82rem;width:100%;border-collapse:collapse}}
      .diff td{{padding:2px 6px;vertical-align:top;white-space:pre-wrap;word-break:break-word}}
      .diff_header{{background:#f0f0f0;font-weight:bold}}
      td.diff_next{{display:none}}
      .diff_add{{background:#d4edda}}.diff_chg{{background:#fff3cd}}.diff_sub{{background:#f8d7da}}
    </style>{diff_html}""", unsafe_allow_html=True)


# ---------------------------------------------------------
# BADGE HELPERS
# ---------------------------------------------------------

BADGE_STYLES = {
    "high":   ("🟢", "#1a7a3c", "#d4edda"),
    "medium": ("🟡", "#856404", "#fff3cd"),
    "low":    ("🔴", "#721c24", "#f8d7da"),
}

def confidence_badge(level: str) -> str:
    level = (level or "low").lower()
    icon, color, bg = BADGE_STYLES.get(level, BADGE_STYLES["low"])
    return (f'<span style="background:{bg};color:{color};padding:2px 8px;'
            f'border-radius:12px;font-size:.75rem;font-weight:600;">{icon} {level.upper()}</span>')

def intent_score_badge(score: int) -> str:
    if score >= 8:   color, bg, label = "#1a7a3c", "#d4edda", "🔥 HIGH"
    elif score >= 5: color, bg, label = "#856404", "#fff3cd", "⚡ MEDIUM"
    else:            color, bg, label = "#721c24", "#f8d7da", "❄️ LOW"
    return (f'<span style="background:{bg};color:{color};padding:3px 10px;'
            f'border-radius:12px;font-size:.85rem;font-weight:700;">{label} &nbsp; {score}/10</span>')


# ---------------------------------------------------------
# EMAIL EXPANDER WIDGET
# ---------------------------------------------------------

def render_email_expander(profile: dict, key_prefix: str):
    company = profile.get("company_name", "unknown")
    with st.expander("✉️ Outreach & Email Drafts"):

        tab_single, tab_sequence, tab_linkedin, tab_objections = st.tabs(
            ["📧 Cold Email", "🔁 Follow-up Sequence", "💼 LinkedIn DM", "🛡 Objection Prep"]
        )

        # ── Tab 1: Cold Email ─────────────────────────────
        with tab_single:
            st.markdown("#### ✏️ Generate Draft")
            tone = st.radio("Tone", ["formal","friendly","bold"], horizontal=True, key=f"tone_{key_prefix}")
            if st.button("Generate Email", key=f"gen_email_{key_prefix}"):
                with st.spinner("Writing…"):
                    try:
                        raw = generate_email_draft(profile, tone)
                        subj, body = parse_email(raw)
                        st.session_state[f"draft_subject_{key_prefix}"] = subj
                        st.session_state[f"draft_body_{key_prefix}"]    = body
                    except Exception as e:
                        st.error(f"Failed: {e}")

            subj_val = st.session_state.get(f"draft_subject_{key_prefix}", "")
            body_val = st.session_state.get(f"draft_body_{key_prefix}", "")
            if body_val:
                if subj_val: st.markdown(f"**Subject:** {subj_val}")
                edited = st.text_area("Body", value=body_val, height=200, key=f"email_body_{key_prefix}")
                if st.button("💾 Save", key=f"save_draft_{key_prefix}"):
                    save_email_draft(company, subj_val, edited, tone)
                    st.success("Draft saved")
                    st.rerun()

            saved_rows = load_drafts_for_company(company)
            if saved_rows:
                st.divider()
                st.markdown("#### 📂 Saved Drafts")
                opts = {f"#{r['id']} · {r['tone']} · {'⚡' if r['is_enhanced'] else '🖊'} · {r['created_at']}": r
                        for r in saved_rows}
                sel_label = st.selectbox("Base draft", list(opts.keys()), key=f"base_select_{key_prefix}")
                sel       = dict(opts[sel_label])
                with st.container(border=True):
                    st.caption(f"Subject: {sel.get('subject','—')}")
                    st.text(sel["body"])
                col_del, _ = st.columns([1,4])
                with col_del:
                    if st.button("🗑 Delete", key=f"del_draft_{key_prefix}"):
                        delete_email_draft(sel["id"]); st.rerun()
                st.divider()
                st.markdown("#### 🚀 Enhance from Base")
                enh_tone = st.radio("Tone", ["formal","friendly","bold"], horizontal=True, key=f"enh_tone_{key_prefix}")
                instr = st.text_area("Instructions", placeholder="e.g. More urgent opening, add a stat…",
                                     height=80, key=f"enh_instr_{key_prefix}")
                if st.button("✨ Enhance", key=f"enh_btn_{key_prefix}") and instr.strip():
                    with st.spinner("Enhancing…"):
                        try:
                            raw_enh = generate_enhanced_draft(profile, sel.get("subject",""), sel["body"], instr, enh_tone)
                            es, eb  = parse_email(raw_enh)
                            st.session_state[f"enh_subject_{key_prefix}"] = es
                            st.session_state[f"enh_body_{key_prefix}"]    = eb
                            st.session_state[f"enh_base_{key_prefix}"]    = sel["id"]
                        except Exception as e:
                            st.error(f"Failed: {e}")
                eb_val = st.session_state.get(f"enh_body_{key_prefix}", "")
                es_val = st.session_state.get(f"enh_subject_{key_prefix}", "")
                if eb_val:
                    st.markdown("##### 🔍 Diff")
                    render_diff(sel["body"], eb_val)
                    st.markdown(f"**Subject:** {es_val}")
                    eb_edit = st.text_area("Enhanced body", value=eb_val, height=180, key=f"enh_edit_{key_prefix}")
                    if st.button("💾 Save enhanced", key=f"save_enh_{key_prefix}"):
                        save_email_draft(company, es_val, eb_edit, enh_tone, is_enhanced=True,
                                         base_draft_id=st.session_state.get(f"enh_base_{key_prefix}"))
                        st.success("Enhanced draft saved!")
                        for k in (f"enh_body_{key_prefix}", f"enh_subject_{key_prefix}", f"enh_base_{key_prefix}"):
                            st.session_state.pop(k, None)
                        st.rerun()

        # ── Tab 2: Follow-up Sequence ─────────────────────
        with tab_sequence:
            st.markdown("#### 🔁 3-Touch Follow-up Sequence")
            seq_tone = st.radio("Tone", ["formal","friendly","bold"], horizontal=True, key=f"seq_tone_{key_prefix}")
            if st.button("Generate Sequence", key=f"gen_seq_{key_prefix}"):
                with st.spinner("Writing 3-email sequence…"):
                    try:
                        seq = generate_followup_sequence(profile, seq_tone)
                        st.session_state[f"sequence_{key_prefix}"] = seq
                    except Exception as e:
                        st.error(f"Failed: {e}")
            seq = st.session_state.get(f"sequence_{key_prefix}", [])
            if seq:
                for email in seq:
                    with st.container(border=True):
                        st.markdown(f"**Touch {email.get('touch')} — {email.get('send_gap')}**")
                        st.markdown(f"*Subject: {email.get('subject','')}*")
                        st.text_area("Body", value=email.get("body",""), height=140,
                                     key=f"seq_body_{key_prefix}_{email.get('touch')}")
                        if st.button(f"💾 Save touch {email.get('touch')}", key=f"save_seq_{key_prefix}_{email.get('touch')}"):
                            save_email_draft(company, email.get("subject",""), email.get("body",""),
                                             seq_tone + f"_touch{email.get('touch')}")
                            st.success(f"Touch {email.get('touch')} saved")

        # ── Tab 3: LinkedIn DM ────────────────────────────
        with tab_linkedin:
            st.markdown("#### 💼 LinkedIn Connection Note")
            st.caption("Max 300 characters — optimised for LinkedIn DMs.")
            if st.button("Generate LinkedIn Message", key=f"gen_li_{key_prefix}"):
                with st.spinner("Writing…"):
                    try:
                        msg = generate_linkedin_message(profile)
                        st.session_state[f"li_msg_{key_prefix}"] = msg
                    except Exception as e:
                        st.error(f"Failed: {e}")
            li_msg = st.session_state.get(f"li_msg_{key_prefix}", "")
            if li_msg:
                edited_li = st.text_area("Message", value=li_msg, height=100, key=f"li_edit_{key_prefix}")
                char_count = len(edited_li)
                color = "green" if char_count <= 300 else "red"
                st.markdown(f'<span style="color:{color};font-size:.8rem;">{char_count}/300 characters</span>',
                            unsafe_allow_html=True)
                if st.button("💾 Save LinkedIn message", key=f"save_li_{key_prefix}"):
                    save_email_draft(company, "LinkedIn DM", edited_li, "linkedin")
                    st.success("Saved")

        # ── Tab 4: Objection Prep ─────────────────────────
        with tab_objections:
            st.markdown("#### 🛡 Objection Handling Prep")
            if st.button("Generate Objections & Rebuttals", key=f"gen_obj_{key_prefix}"):
                with st.spinner("Generating…"):
                    try:
                        objs = generate_objection_prep(profile)
                        st.session_state[f"objections_{key_prefix}"] = objs
                    except Exception as e:
                        st.error(f"Failed: {e}")
            objs = st.session_state.get(f"objections_{key_prefix}", [])
            if objs:
                for i, item in enumerate(objs, 1):
                    with st.container(border=True):
                        st.markdown(f"**Objection {i}:** {item.get('objection','')}")
                        st.markdown(f"**Rebuttal:** {item.get('rebuttal','')}")


# ---------------------------------------------------------
# NOTES WIDGET  (Step 3)
# ---------------------------------------------------------

def render_notes_expander(company_name: str, key_prefix: str):
    with st.expander("📝 Notes & Activity Log"):
        note_input = st.text_area("Add a note", placeholder="e.g. Called VP Sales, follow up in 2 weeks…",
                                  height=80, key=f"note_input_{key_prefix}")
        if st.button("Add Note", key=f"add_note_{key_prefix}") and note_input.strip():
            save_note(company_name, note_input.strip())
            st.success("Note saved")
            st.rerun()
        notes = load_notes(company_name)
        if notes:
            st.divider()
            for n in notes:
                col_txt, col_del = st.columns([6, 1])
                with col_txt:
                    st.markdown(f"**{n['created_at']}** — {n['note']}")
                with col_del:
                    if st.button("🗑", key=f"del_note_{n['id']}"):
                        delete_note(n["id"]); st.rerun()
        else:
            st.caption("No notes yet.")


# ---------------------------------------------------------
# PROFILE RENDERERS
# ---------------------------------------------------------

def render_overview_tab(p: dict, key_prefix: str):
    conf = p.get("confidence", {})
    for field, label in [
        ("core_product", "Core Product"),
        ("recent_news",  "Recent News"),
    ]:
        st.markdown(f"**{label}** &nbsp; {confidence_badge(conf.get(field,'low'))}",
                    unsafe_allow_html=True)
        st.markdown(p.get(field, ""))

    st.markdown(f"**🎯 Pain Points** &nbsp; {confidence_badge(conf.get('pain_points','low'))}",
                unsafe_allow_html=True)
    for pt in p.get("pain_points", []):
        st.markdown(f"- {pt}")

    st.markdown(f"**💡 Sales Pitch Angle** &nbsp; {confidence_badge(conf.get('pitch_angle','low'))}",
                unsafe_allow_html=True)
    st.markdown(p.get("pitch_angle", ""))

    render_email_expander(p, key_prefix)
    render_notes_expander(p.get("company_name",""), key_prefix)


def render_funding_tab(funding: dict, conf: dict):
    st.markdown(f"**💰 Total Raised** &nbsp; {confidence_badge(conf.get('funding_info','low'))}",
                unsafe_allow_html=True)
    st.markdown(funding.get("total_raised", "Unknown"))
    st.markdown("**📅 Last Round**"); st.markdown(funding.get("last_round", "Unknown"))
    st.markdown("**🏦 Key Investors**"); st.markdown(funding.get("key_investors", "Unknown"))
    st.markdown("**📊 Revenue / Headcount Signals**"); st.markdown(funding.get("revenue_signals", "Unknown"))


def render_jobs_tab(job_signals: dict):
    """Step 1 — Job Signals tab."""
    if not job_signals or job_signals.get("hiring_themes") == "Unknown":
        st.info("No job signal data. Re-research with Deep depth to populate this tab.")
        return
    open_roles = job_signals.get("open_roles", [])
    if open_roles:
        st.markdown("**👥 Open Roles Detected**")
        for role in open_roles:
            st.markdown(f"- {role}")
    st.markdown("**📈 Hiring Themes**")
    st.markdown(job_signals.get("hiring_themes", "Unknown"))
    st.markdown("**🔢 Headcount Signal**")
    st.markdown(job_signals.get("headcount_signal", "Unknown"))
    st.markdown("**💡 Pitch Implication**")
    st.markdown(job_signals.get("pitch_implication", "Unknown"))


def render_stack_tab(tech_stack: dict):
    """Step 1 — Tech Stack tab."""
    if not tech_stack or tech_stack.get("stack_summary") == "Unknown":
        st.info("No tech stack data. Re-research with Deep depth to populate this tab.")
        return
    tools = tech_stack.get("tools_identified", [])
    if tools:
        st.markdown("**🛠 Tools & Platforms Identified**")
        cols = st.columns(min(len(tools), 4))
        for i, tool_name in enumerate(tools):
            cols[i % 4].markdown(
                f'<span style="background:#e9ecef;padding:3px 8px;border-radius:8px;'
                f'font-size:.8rem;">{tool_name}</span>', unsafe_allow_html=True)
    st.markdown("**📋 Stack Summary**")
    st.markdown(tech_stack.get("stack_summary", "Unknown"))
    st.markdown("**💡 Pitch Implication**")
    st.markdown(tech_stack.get("pitch_implication", "Unknown"))


def render_competitors_tab(competitors: list, key_prefix: str):
    if not competitors:
        st.info("No competitor data found.")
        return
    for i, comp in enumerate(competitors):
        with st.container(border=True):
            st.markdown(f"### 🏢 {comp.get('company_name','Unknown')}")
            conf = comp.get("confidence", {})
            for field, label in [("core_product","Core Product"),("recent_news","Recent News")]:
                st.markdown(f"**{label}** &nbsp; {confidence_badge(conf.get(field,'low'))}",
                            unsafe_allow_html=True)
                st.markdown(comp.get(field, ""))
            st.markdown(f"**🎯 Pain Points** &nbsp; {confidence_badge(conf.get('pain_points','low'))}",
                        unsafe_allow_html=True)
            for pt in comp.get("pain_points", []):
                st.markdown(f"- {pt}")
            st.markdown(f"**💡 Differentiation** &nbsp; {confidence_badge(conf.get('pitch_angle','low'))}",
                        unsafe_allow_html=True)
            st.markdown(comp.get("pitch_angle", ""))
            render_email_expander(comp, key_prefix=f"{key_prefix}_comp{i}")


def render_persona_followup_prompt(p: dict, key_prefix: str = ""):
    current_provider = get_llm_provider()
    current_model_name = get_llm_model(current_provider)
    current_preset = get_persona_preset_for_model(current_provider, current_model_name)
    followup_options = [preset for preset in PERSONA_PRESETS.keys() if preset != current_preset]
    if not followup_options:
        return

    followup_state_key = f"persona_followup_result_{key_prefix}"
    followup_meta_key = f"persona_followup_meta_{key_prefix}"
    with st.container(border=True):
        st.markdown("#### 🔁 Analyze further with a different persona")
        st.caption("Pick a different persona for a second pass on this analysis.")
        selected_followup = st.selectbox(
            "Follow-up persona",
            options=followup_options,
            format_func=lambda key: PERSONA_PRESET_LABELS.get(key, key),
            key=f"persona_followup_select_{key_prefix}",
            label_visibility="collapsed",
        )
        if st.button("Run follow-up analysis", key=f"persona_followup_run_{key_prefix}", use_container_width=True):
            with st.spinner("Running follow-up analysis…"):
                try:
                    followup_prompt = PERSONA_PRESETS[selected_followup]
                    full_system_prompt = followup_prompt + build_company_memory_block(load_all_profiles())
                    pd_, conf, fi, comps, jobs, stack = enrich_company(
                        p["company_name"],
                        full_system_prompt,
                        p.get("search_depth", st.session_state.search_depth),
                    )
                    followup_profile = {
                        **pd_,
                        "confidence": conf,
                        "funding_info": fi,
                        "competitors": comps,
                        "chat_history": [],
                        "search_depth": p.get("search_depth", st.session_state.search_depth),
                        "job_signals": jobs,
                        "tech_stack": stack,
                        "id": None,
                    }
                    st.session_state[followup_meta_key] = {
                        "persona": selected_followup,
                        "note": get_persona_comparison_note(selected_followup),
                    }
                    st.session_state[followup_state_key] = followup_profile
                    st.success(f"Follow-up analysis complete with {PERSONA_PRESET_LABELS.get(selected_followup, selected_followup)}.")
                except Exception as e:
                    st.error(f"Follow-up analysis failed: {e}")

    followup_profile = st.session_state.get(followup_state_key)
    if followup_profile:
        followup_meta = st.session_state.get(followup_meta_key, {})
        with st.expander("🧭 Follow-up Analysis", expanded=False):
            if followup_meta.get("persona"):
                st.caption(
                    f"Persona: {PERSONA_PRESET_LABELS.get(followup_meta['persona'], followup_meta['persona'])} — {followup_meta.get('note', '')}"
                )
            render_profile_tabs(followup_profile, key_prefix=f"{key_prefix}_followup", include_followup=False)


def render_profile_tabs(p: dict, key_prefix: str = "", include_followup: bool = True):
    depth_badge = {"fast":"⚡ Fast","balanced":"⚖️ Balanced","deep":"🔬 Deep"}.get(p.get("search_depth",""),"")
    st.markdown(f"### 🏢 {p['company_name']}  &nbsp; <small>{depth_badge}</small>",
                unsafe_allow_html=True)

    tab_ov, tab_fi, tab_jobs, tab_stack, tab_comp = st.tabs(
        ["📋 Overview", "💰 Funding", "👥 Job Signals", "🛠 Tech Stack", "🏆 Competitors"]
    )
    kp = key_prefix or p.get("company_name","p").replace(" ","_")
    with tab_ov:   render_overview_tab(p, kp)
    with tab_fi:   render_funding_tab(p.get("funding_info",{}), p.get("confidence",{}))
    with tab_jobs: render_jobs_tab(p.get("job_signals",{}))
    with tab_stack:render_stack_tab(p.get("tech_stack",{}))
    with tab_comp: render_competitors_tab(p.get("competitors",[]), kp)
    if include_followup:
        render_persona_followup_prompt(p, kp)


def profile_to_markdown(p: dict) -> str:
    conf = p.get("confidence", {})
    bt   = {"high":"[HIGH]","medium":"[MEDIUM]","low":"[LOW]"}
    def b(f): return bt.get((conf.get(f) or "low").lower(),"[LOW]")
    md  = f"### 🏢 {p['company_name']}\n\n"
    md += f"**Core Product** {b('core_product')}\n{p['core_product']}\n\n"
    md += f"**Recent News** {b('recent_news')}\n{p['recent_news']}\n\n"
    md += f"### 🎯 Pain Points {b('pain_points')}\n\n"
    for pt in p["pain_points"]: md += f"- {pt}\n"
    md += f"\n### 💡 Pitch Angle {b('pitch_angle')}\n\n{p['pitch_angle']}\n"
    fi = p.get("funding_info",{})
    if fi:
        md += f"\n### 💰 Funding\n- Raised: {fi.get('total_raised','?')}\n"
        md += f"- Last round: {fi.get('last_round','?')}\n"
    js = p.get("job_signals",{})
    if js and js.get("hiring_themes","Unknown") != "Unknown":
        md += f"\n### 👥 Hiring\n{js.get('hiring_themes','')}\n"
    ts = p.get("tech_stack",{})
    if ts and ts.get("tools_identified"):
        md += f"\n### 🛠 Stack\n{', '.join(ts['tools_identified'])}\n"
    return md


# ---------------------------------------------------------
# EXPORT HELPERS
# ---------------------------------------------------------

def profiles_to_csv(profiles: list) -> bytes:
    output = io.StringIO()
    fieldnames = [
        "company_name","core_product","recent_news","pain_points","pitch_angle",
        "conf_core","conf_news","conf_pain","conf_pitch",
        "funding_raised","funding_round","funding_investors","funding_revenue","conf_funding",
        "jobs_roles","jobs_themes","jobs_headcount","jobs_pitch",
        "stack_tools","stack_summary","stack_pitch",
        "competitors","search_depth","created_at",
    ]
    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()
    for p in profiles:
        conf = p.get("confidence",{})
        fi   = p.get("funding_info",{})
        js   = p.get("job_signals",{})
        ts   = p.get("tech_stack",{})
        writer.writerow({
            "company_name":      p["company_name"],
            "core_product":      p["core_product"],
            "recent_news":       p["recent_news"],
            "pain_points":       "; ".join(p["pain_points"]),
            "pitch_angle":       p["pitch_angle"],
            "conf_core":         conf.get("core_product",""),
            "conf_news":         conf.get("recent_news",""),
            "conf_pain":         conf.get("pain_points",""),
            "conf_pitch":        conf.get("pitch_angle",""),
            "funding_raised":    fi.get("total_raised",""),
            "funding_round":     fi.get("last_round",""),
            "funding_investors": fi.get("key_investors",""),
            "funding_revenue":   fi.get("revenue_signals",""),
            "conf_funding":      conf.get("funding_info",""),
            "jobs_roles":        "; ".join(js.get("open_roles",[])),
            "jobs_themes":       js.get("hiring_themes",""),
            "jobs_headcount":    js.get("headcount_signal",""),
            "jobs_pitch":        js.get("pitch_implication",""),
            "stack_tools":       "; ".join(ts.get("tools_identified",[])),
            "stack_summary":     ts.get("stack_summary",""),
            "stack_pitch":       ts.get("pitch_implication",""),
            "competitors":       "; ".join(c.get("company_name","") for c in p.get("competitors",[])),
            "search_depth":      p.get("search_depth",""),
            "created_at":        p.get("created_at",""),
        })
    return output.getvalue().encode("utf-8")


CONF_FILL = {
    "high":   PatternFill("solid", fgColor="C6EFCE"),
    "medium": PatternFill("solid", fgColor="FFEB9C"),
    "low":    PatternFill("solid", fgColor="FFC7CE"),
}

def profiles_to_xlsx(profiles: list) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Enriched Leads"
    headers = [
        "Company","Core Product","Recent News","Pain Points","Pitch Angle",
        "Conf:Core","Conf:News","Conf:Pain","Conf:Pitch",
        "Funding Raised","Last Round","Investors","Revenue Signals","Conf:Funding",
        "Open Roles","Hiring Themes","Headcount","Jobs Pitch",
        "Tech Tools","Stack Summary","Stack Pitch",
        "Competitors","Depth","Researched At",
    ]
    hf = PatternFill("solid", fgColor="1F4E79")
    hfont = Font(bold=True, color="FFFFFF")
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill, c.font, c.alignment = hf, hfont, Alignment(wrap_text=True)

    for ri, p in enumerate(profiles, 2):
        conf = p.get("confidence",{})
        fi   = p.get("funding_info",{})
        js   = p.get("job_signals",{})
        ts   = p.get("tech_stack",{})
        vals = [
            p["company_name"], p["core_product"], p["recent_news"],
            "; ".join(p["pain_points"]), p["pitch_angle"],
            conf.get("core_product",""), conf.get("recent_news",""),
            conf.get("pain_points",""),  conf.get("pitch_angle",""),
            fi.get("total_raised",""),   fi.get("last_round",""),
            fi.get("key_investors",""),  fi.get("revenue_signals",""), conf.get("funding_info",""),
            "; ".join(js.get("open_roles",[])), js.get("hiring_themes",""),
            js.get("headcount_signal",""), js.get("pitch_implication",""),
            "; ".join(ts.get("tools_identified",[])), ts.get("stack_summary",""),
            ts.get("pitch_implication",""),
            "; ".join(c.get("company_name","") for c in p.get("competitors",[])),
            p.get("search_depth",""), p.get("created_at",""),
        ]
        for ci, v in enumerate(vals, 1):
            ws.cell(row=ri, column=ci, value=v).alignment = Alignment(wrap_text=True, vertical="top")
        for ci, ck in zip([6,7,8,9,14],["core_product","recent_news","pain_points","pitch_angle","funding_info"]):
            ws.cell(row=ri, column=ci).fill = CONF_FILL.get((conf.get(ck) or "low").lower(), CONF_FILL["low"])

    for i, w in enumerate([22,38,38,46,46,10,10,10,10,18,18,26,26,10,
                            30,30,20,30,30,30,30,28,10,16], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------
# INTENT PRIORITY DASHBOARD
# ---------------------------------------------------------

def render_priority_dashboard(all_dicts: list):
    st.title("🎯 Priority Dashboard")
    st.caption("Companies ranked by intent score. Scores cached — click Refresh to recompute.")
    if not all_dicts:
        st.info("No profiles yet.")
        return

    col_a, col_b, _ = st.columns([2,2,6])
    with col_a:
        score_all = st.button("🔄 Score / Refresh All", use_container_width=True)
    with col_b:
        filt = st.selectbox("Filter", ["All","🔥 High (8–10)","⚡ Medium (5–7)","❄️ Low (1–4)"],
                            label_visibility="collapsed")
    st.divider()

    if score_all:
        prog = st.progress(0, text="Scoring…")
        for i, p in enumerate(all_dicts):
            prog.progress(i / len(all_dicts), text=f"Scoring {p['company_name']}…")
            sd = compute_intent_score(p)
            save_intent_score(p["id"], sd)
            p["intent_score"] = sd
        prog.progress(1.0, text="✅ Done")
        st.rerun()

    scored = sorted(all_dicts,
                    key=lambda x: (x["intent_score"] or {}).get("score", -1) if x.get("intent_score") else -1,
                    reverse=True)

    def passes(item):
        s = item.get("intent_score")
        if filt == "All": return True
        if not s: return False
        v = s["score"]
        if filt.startswith("🔥"): return v >= 8
        if filt.startswith("⚡"): return 5 <= v <= 7
        return v <= 4

    visible = [x for x in scored if passes(x)]
    if not visible:
        st.info("No companies match the filter.")
        return

    for item in visible:
        intent = item.get("intent_score")
        with st.container(border=True):
            col_n, col_b2, col_btn = st.columns([4,3,2])
            with col_n:
                di = {"fast":"⚡","balanced":"⚖️","deep":"🔬"}.get(item.get("search_depth",""),"")
                st.markdown(f"### {di} {item['company_name']}")
                st.caption(item.get("created_at",""))
            with col_b2:
                if intent:
                    st.markdown(intent_score_badge(intent["score"]), unsafe_allow_html=True)
                else:
                    st.caption("Not scored")
            with col_btn:
                if st.button("Score now", key=f"score_{item['id']}"):
                    with st.spinner("Scoring…"):
                        sd = compute_intent_score(item)
                        save_intent_score(item["id"], sd)
                    st.rerun()

            if intent:
                rp = intent.get("rule_score", 0)
                st.markdown(
                    f'<div style="background:#e9ecef;border-radius:6px;height:6px;margin:4px 0 8px;">'
                    f'<div style="background:#0d6efd;width:{rp}%;height:6px;border-radius:6px;"></div></div>'
                    f'<span style="font-size:.72rem;color:#6c757d;">Rule signal strength: {rp}/100</span>',
                    unsafe_allow_html=True)
                if intent.get("signals"):
                    with st.expander("📡 Triggered signals"):
                        for sig in intent["signals"]: st.markdown(f"- ✅ {sig}")
                col1, col2 = st.columns(2)
                with col1:
                    st.markdown("**🧠 Reasoning**")
                    st.markdown(intent.get("reasoning","—"))
                with col2:
                    st.markdown("**📋 Recommended Action**")
                    st.markdown(intent.get("recommended_action","—"))
                    st.markdown("**🕐 Best Time to Reach**")
                    st.markdown(intent.get("best_time_to_reach","—"))
                if st.button(f"Open {item['company_name']} →", key=f"jump_{item['id']}"):
                    st.session_state.viewing_profile_id = item["id"]
                    st.session_state.current_page = "research"
                    st.rerun()


# ---------------------------------------------------------
# TRIGGER ALERT / RE-RESEARCH PAGE  (Step 1)
# ---------------------------------------------------------

def render_trigger_alerts_page(all_dicts: list):
    st.title("🔔 Trigger Alerts — Re-research")
    st.caption("Re-run the full research pipeline on a saved company and see what changed.")

    if not all_dicts:
        st.info("No profiles yet.")
        return

    company_names = [p["company_name"] for p in all_dicts]
    selected_name = st.selectbox("Select company to re-research", company_names)
    old_profile   = next(p for p in all_dicts if p["company_name"] == selected_name)

    col_depth, col_btn, _ = st.columns([2, 2, 6])
    with col_depth:
        depth = st.radio("Depth", ["fast","balanced","deep"],
                         format_func=lambda x: {"fast":"⚡","balanced":"⚖️","deep":"🔬"}[x]+" "+x.title(),
                         horizontal=True, key="alert_depth")
    with col_btn:
        run = st.button("🔄 Re-research Now", type="primary", use_container_width=True)

    if run:
        user_sys = get_current_system_prompt()
        full_sys = user_sys + build_company_memory_block(load_all_profiles())
        with st.spinner(f"Re-researching {selected_name}…"):
            try:
                pd_, conf, fi, comps, jobs, stack = enrich_company(
                    selected_name, full_sys, depth
                )
                new_profile = {**pd_, "confidence": conf, "funding_info": fi,
                               "competitors": comps, "job_signals": jobs,
                               "tech_stack": stack, "search_depth": depth}

                changes = diff_profiles(old_profile, new_profile)

                st.subheader("📊 What Changed")
                for change in changes:
                    if change == "No significant changes detected.":
                        st.info(change)
                    else:
                        st.success(change)

                # Save as a new profile entry
                save_profile(pd_, conf, fi, comps, [], depth, jobs, stack)
                st.info("New profile saved to history.")

                with st.expander("🔍 Full New Profile"):
                    render_profile_tabs(new_profile, key_prefix=f"alert_{selected_name}")

                st.rerun()

            except Exception as e:
                st.error(f"Re-research failed: {e}")


# ---------------------------------------------------------
# BACKUP & RESTORE PAGE  (Step 3)
# ---------------------------------------------------------

def render_backup_page():
    st.title("🗄 DB Backup & Restore")

    st.markdown("### ⬇️ Create Backup")
    if st.button("Create backup now", use_container_width=False):
        dst = create_backup()
        st.success(f"Backup saved: `{dst}`")

    st.markdown("### 📦 Download Backup")
    if DB_PATH.exists():
        st.download_button(
            "⬇️ Download current DB",
            data=DB_PATH.read_bytes(),
            file_name=f"enrichment_profiles_{datetime.now().strftime('%Y%m%d_%H%M')}.db",
            mime="application/octet-stream",
            use_container_width=False,
        )

    backups = list_backups()
    if backups:
        st.markdown("### 🕐 Local Backups")
        for bp in backups[:10]:
            col_name, col_dl = st.columns([5, 1])
            with col_name:
                st.caption(str(bp.name))
            with col_dl:
                st.download_button(
                    "⬇️", data=bp.read_bytes(),
                    file_name=bp.name,
                    mime="application/octet-stream",
                    key=f"dl_{bp.name}",
                )

    st.divider()
    st.markdown("### ⬆️ Restore from File")
    st.warning("⚠️ Restoring will **replace** the current database. This cannot be undone.")
    uploaded = st.file_uploader("Upload a .db backup file", type=["db"])
    if uploaded:
        if st.button("🔁 Restore this backup", type="primary"):
            create_backup()   # auto-backup before restore
            restore_backup(uploaded.read())
            st.success("Database restored. Please restart the app.")


# ---------------------------------------------------------
# LANGSMITH DASHBOARD PAGE
# ---------------------------------------------------------

@st.cache_data(ttl=60)
def fetch_langsmith_stats(project: str) -> dict:
    """Pull run stats from LangSmith API. Cached for 60s."""
    try:
        client = LangSmithClient()
        runs   = list(client.list_runs(
            project_name=project,
            execution_order=1,   # top-level runs only
            limit=200,
        ))
        if not runs:
            return {"error": None, "runs": 0, "total_tokens": 0,
                    "total_cost": 0.0, "errors": 0, "ops": {}}

        total_tokens = sum(getattr(r, "total_tokens", 0) or 0 for r in runs)
        total_cost   = sum(float(getattr(r, "total_cost", 0) or 0) for r in runs)
        error_count  = sum(1 for r in runs if getattr(r, "error", None))

        # Breakdown by operation name (traceable function name)
        ops: dict[str, dict] = {}
        for r in runs:
            name = r.name or "unknown"
            if name not in ops:
                ops[name] = {"count": 0, "errors": 0, "tokens": 0}
            ops[name]["count"]  += 1
            ops[name]["errors"] += 1 if getattr(r, "error", None) else 0
            ops[name]["tokens"] += getattr(r, "total_tokens", 0) or 0

        return {
            "error":        None,
            "runs":         len(runs),
            "total_tokens": total_tokens,
            "total_cost":   total_cost,
            "errors":       error_count,
            "ops":          ops,
        }
    except Exception as e:
        return {"error": str(e), "runs": 0, "total_tokens": 0,
                "total_cost": 0.0, "errors": 0, "ops": {}}


def render_langsmith_page():
    st.title("🔭 LangSmith Observability")

    tracing_on = os.environ.get("LANGSMITH_TRACING", "").lower() == "true"
    api_key    = os.environ.get("LANGSMITH_API_KEY", "")
    project    = os.environ.get("LANGSMITH_PROJECT", LANGSMITH_PROJECT)

    # Status banner
    if tracing_on and api_key:
        st.success(f"✅ Tracing **active** — project: `{project}`")
    else:
        st.warning(
            "⚠️ Tracing is **off**. Add these to your `.env` file to enable:\n\n"
            "```\nLANGSMITH_TRACING=true\n"
            "LANGSMITH_API_KEY=<your key>\n"
            "LANGSMITH_PROJECT=sales-enrichment-agent\n```"
        )
        return

    # Project link
    st.markdown(
        f"[🌐 Open LangSmith Dashboard](https://smith.langchain.com/projects/{project})",
        unsafe_allow_html=False,
    )

    col_refresh, _ = st.columns([2, 8])
    with col_refresh:
        if st.button("🔄 Refresh Stats"):
            fetch_langsmith_stats.clear()
            st.rerun()

    st.divider()

    stats = fetch_langsmith_stats(project)

    if stats.get("error"):
        st.error(f"Could not fetch stats: {stats['error']}")
        st.caption("Check that your LANGSMITH_API_KEY is correct and has read access.")
        return

    # KPI row
    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Total Runs",    stats["runs"])
    col2.metric("Total Tokens",  f"{stats['total_tokens']:,}")
    col3.metric("Est. Cost",     f"${stats['total_cost']:.4f}")
    col4.metric("Errors",        stats["errors"],
                delta=f"{(stats['errors']/max(stats['runs'],1)*100):.1f}% error rate",
                delta_color="inverse")

    st.divider()
    st.markdown("### 📊 Runs by Operation")

    ops = stats.get("ops", {})
    if ops:
        # Sort by count desc
        sorted_ops = sorted(ops.items(), key=lambda x: x[1]["count"], reverse=True)
        for op_name, op_stats in sorted_ops:
            err_rate = op_stats["errors"] / max(op_stats["count"], 1) * 100
            col_n, col_c, col_e, col_t = st.columns([4, 2, 2, 2])
            col_n.markdown(f"**{op_name}**")
            col_c.metric("Runs",   op_stats["count"], label_visibility="collapsed")
            col_e.metric("Errors", op_stats["errors"],
                         delta=f"{err_rate:.0f}%",
                         delta_color="inverse" if op_stats["errors"] else "off",
                         label_visibility="collapsed")
            col_t.metric("Tokens", f"{op_stats['tokens']:,}", label_visibility="collapsed")
    else:
        st.info("No run data yet. Start researching companies to generate traces.")

    st.divider()
    st.markdown("### 🧵 Current Session")
    st.caption(f"Session ID: `{st.session_state.get('session_id','—')}`")
    st.markdown(
        "All traces in this session are tagged with the above ID. "
        "Filter by `session_id` in LangSmith to see only this session's runs."
    )

    # Show .env guidance
    with st.expander("⚙️ Required .env variables"):
        st.code(
            "LANGSMITH_TRACING=true\n"
            f"LANGSMITH_API_KEY=<your key>\n"
            f"LANGSMITH_PROJECT={project}\n"
            "OPENAI_API_KEY=<your key>",
            language="bash",
        )


# ---------------------------------------------------------
# SESSION STATE
# ---------------------------------------------------------

if "messages"           not in st.session_state: st.session_state.messages           = []
if "viewing_profile_id" not in st.session_state: st.session_state.viewing_profile_id = None
if "editing_prompt"     not in st.session_state: st.session_state.editing_prompt     = False
if "bulk_mode"          not in st.session_state: st.session_state.bulk_mode          = False
if "bulk_results"       not in st.session_state: st.session_state.bulk_results       = []
if "export_selection"   not in st.session_state: st.session_state.export_selection   = "all"
if "last_profile"       not in st.session_state: st.session_state.last_profile       = None
if "search_depth"       not in st.session_state: st.session_state.search_depth       = "balanced"
if "current_page"       not in st.session_state: st.session_state.current_page       = "research"
if "conversation_temperature" not in st.session_state: st.session_state.conversation_temperature = 0.7


def switch_to_chat():
    st.session_state.viewing_profile_id = None
    st.session_state.conversation_temperature = 0.7


# ---------------------------------------------------------
# SIDEBAR
# ---------------------------------------------------------

with st.sidebar:

    # ── NAV ───────────────────────────────────────────────
    st.markdown("## 🧭 Navigation")
    pages = [
        ("research",   "🔬 Research"),
        ("dashboard",  "🎯 Dashboard"),
        ("alerts",     "🔔 Alerts"),
        ("langsmith",  "🔭 LangSmith"),
        ("backup",     "🗄 Backup"),
    ]
    for page_id, page_label in pages:
        active = st.session_state.current_page == page_id
        if st.button(
            f"{'▶ ' if active else ''}{page_label}",
            use_container_width=True,
            type="primary" if active else "secondary",
            key=f"nav_{page_id}",
        ):
            st.session_state.current_page = page_id
            st.rerun()

    st.divider()

    # ── SETTINGS ──────────────────────────────────────────
    with st.expander("🧬 Agent Persona", expanded=st.session_state.editing_prompt):
        current_provider = get_llm_provider()
        current_model_name = get_llm_model(current_provider)
        current_preset = get_persona_preset_for_model(current_provider, current_model_name)
        current_prompt = get_current_system_prompt(current_provider, current_model_name)
        preset_options = list(PERSONA_PRESETS.keys())
        selected_preset = st.selectbox(
            "Persona preset",
            options=preset_options,
            index=preset_options.index(current_preset),
            format_func=lambda key: PERSONA_PRESET_LABELS.get(key, key),
            label_visibility="collapsed",
        )
        if selected_preset != current_preset:
            save_persona_preset_for_model(current_provider, current_model_name, selected_preset)
            reset_system_prompt_for_model(current_provider, current_model_name)
            st.session_state.editing_prompt = False; st.rerun()

        st.caption(f"Applies to {current_provider} / {current_model_name}")
        if st.session_state.editing_prompt:
            new_prompt = st.text_area("Prompt", value=current_prompt, height=220, label_visibility="collapsed")
            c1, c2 = st.columns(2)
            with c1:
                if st.button("💾 Save", use_container_width=True):
                    save_system_prompt_for_model(current_provider, current_model_name, new_prompt.strip())
                    st.session_state.editing_prompt = False; st.rerun()
            with c2:
                if st.button("✕ Cancel", use_container_width=True):
                    st.session_state.editing_prompt = False; st.rerun()
        else:
            st.caption(current_prompt[:180] + ("…" if len(current_prompt) > 180 else ""))
            if st.button("✏️ Edit Persona", use_container_width=True):
                st.session_state.editing_prompt = True; st.rerun()
        if st.button("↺ Reset to Default", use_container_width=True):
            save_persona_preset_for_model(current_provider, current_model_name, DEFAULT_PERSONA_PRESET)
            reset_system_prompt_for_model(current_provider, current_model_name)
            st.session_state.editing_prompt = False; st.rerun()

    with st.expander("⚙️ LLM Settings", expanded=False):
        provider_labels = {
            "openrouter": "OpenRouter",
            "openai": "OpenAI",
            "perplexity": "Perplexity",
            "ollama": "Ollama Cloud (Free Tier)",
        }
        current_provider = get_llm_provider()
        selected_provider = st.selectbox(
            "Provider",
            options=list(LLM_PROVIDERS),
            index=list(LLM_PROVIDERS).index(current_provider),
            format_func=lambda x: provider_labels.get(x, x),
            label_visibility="collapsed",
        )
        if selected_provider != current_provider:
            save_setting("llm_provider", selected_provider)
            st.rerun()

        if current_provider == "openrouter":
            openrouter_base = load_setting("openrouter_base_url") or get_openrouter_base_url()
            openrouter_base_input = st.text_input(
                "OpenRouter base URL",
                value=openrouter_base,
                placeholder="https://openrouter.ai/api/v1",
                label_visibility="collapsed",
            )
            if openrouter_base_input.strip() != openrouter_base:
                save_setting("openrouter_base_url", openrouter_base_input.strip())
                st.rerun()

            openrouter_key = load_setting("openrouter_api_key") or get_openrouter_api_key()
            openrouter_key_input = st.text_input(
                "OpenRouter API key",
                value=openrouter_key,
                type="password",
                placeholder="sk-or-...",
                label_visibility="collapsed",
            )
            if openrouter_key_input != openrouter_key:
                save_setting("openrouter_api_key", openrouter_key_input)
                st.rerun()

        if current_provider == "perplexity":
            pplx_base = load_setting("pplx_base_url") or get_perplexity_base_url()
            pplx_base_input = st.text_input(
                "Perplexity base URL",
                value=pplx_base,
                placeholder="https://api.perplexity.ai",
                label_visibility="collapsed",
            )
            if pplx_base_input.strip() != pplx_base:
                save_setting("pplx_base_url", pplx_base_input.strip())
                st.rerun()

            pplx_key = load_setting("pplx_api_key")
            pplx_key_input = st.text_input(
                "Perplexity API key",
                value=pplx_key,
                type="password",
                placeholder="pplx-...",
                label_visibility="collapsed",
            )
            if pplx_key_input != pplx_key:
                save_setting("pplx_api_key", pplx_key_input)
                st.rerun()

        if current_provider == "ollama":
            ollama_base = load_setting("ollama_base_url") or get_ollama_base_url()
            ollama_base_input = st.text_input(
                "Ollama Cloud base URL",
                value=ollama_base,
                placeholder="https://ollama.com",
                label_visibility="collapsed",
            )
            if ollama_base_input.strip() != ollama_base:
                save_setting("ollama_base_url", ollama_base_input.strip())
                st.rerun()

            ollama_insecure_ssl = get_ollama_insecure_ssl()
            ollama_insecure_ssl_input = st.checkbox(
                "Bypass SSL certificate verification",
                value=ollama_insecure_ssl,
                help="Enable this if Ollama Cloud returns an expired or untrusted certificate error.",
            )
            if ollama_insecure_ssl_input != ollama_insecure_ssl:
                save_setting("ollama_insecure_ssl", "true" if ollama_insecure_ssl_input else "false")
                st.rerun()

            ollama_key = load_setting("ollama_api_key") or get_ollama_api_key()
            ollama_key_input = st.text_input(
                "Ollama Cloud API key",
                value=ollama_key,
                type="password",
                placeholder="ollama-...",
                label_visibility="collapsed",
            )
            if ollama_key_input != ollama_key:
                save_setting("ollama_api_key", ollama_key_input)
                st.rerun()

            st.caption(f"API key source: {get_ollama_api_key_source()}")
            if st.button("Test Ollama connection", use_container_width=True):
                with st.spinner("Testing Ollama Cloud…"):
                    try:
                        result = test_ollama_connection()
                        st.success(f"Ollama Cloud OK: {result[:200]}")
                    except Exception as e:
                        st.error(f"Ollama Cloud test failed: {e}")

        current_timeout = get_llm_timeout()
        timeout_input = st.number_input(
            "Timeout seconds",
            min_value=10,
            max_value=900,
            step=5,
            value=current_timeout,
            label_visibility="collapsed",
        )
        if int(timeout_input) != current_timeout:
            save_setting("llm_timeout", str(int(timeout_input)))
            st.rerun()

    with st.expander("🤖 AI Model", expanded=False):
        current_provider = selected_provider
        if current_provider == "openrouter":
            available_models = [
                "openai/gpt-oss-120b:free",
                "openrouter/owl-alpha:free",
                "nvidia/nemotron-3-nano-omni-30b-a3b-reasoning:free",
                "nvidia/nemotron-3-super:free",
                "poolside/laguna-xs-2:free",
                "poolside/laguna-m-1:free",
                "deepseek/deepseek-v4-flash:free",
                "moonshotai/kimi-k2.6:free",
                "google/gemma-4-26b-a4b:free",
                "google/gemma-4-31b:free",
                "google/lyria-3-pro-preview:free",
                "google/lyria-3-clip-preview:free",
                "nvidia/nemotron-3-nano-30b-a3b:free",
                "nvidia/nemotron-nano-12b-2-vl:free",
                "qwen/qwen3-next-80b-a3b-instruct:free",
                "nvidia/nemotron-nano-9b-v2:free",
                "openai/gpt-oss-20b:free",
                "z-ai/glm-4.5-air:free",
                "qwen/qwen3-coder-480b-a35b:free",
                "venice/uncensored:free",
                "meta-llama/llama-3.3-70b-instruct:free",
                "meta-llama/llama-3.2-3b-instruct:free",
                "nousresearch/hermes-3-405b-instruct:free",
                "minimax/minimax-m2.5:free",
                "liquid/lfm2.5-1.2b-thinking:free",
                "liquid/lfm2.5-1.2b-instruct:free",
            ]
        elif current_provider == "perplexity":
            available_models = [
                "sonar",
                "sonar-pro",
                "sonar-reasoning",
                "sonar-deep-research",
            ]
        elif current_provider == "ollama":
            available_models = [
                "gpt-oss:120b",
            ]
        else:
            available_models = [
                "gpt-5-nano",
                "gpt-4.1",
                "gpt-4.1-mini",
                "gpt-4o",
                "gpt-4o-mini",
                "o4-mini",
            ]

        current_model = get_llm_model(current_provider)
        if current_model not in available_models:
            available_models.insert(0, current_model)
        selected_model = st.selectbox(
            "Model",
            options=available_models,
            index=available_models.index(current_model),
            label_visibility="collapsed",
        )
        custom_model = st.text_input(
            "Or enter a custom model name",
            placeholder="e.g. openai/gpt-oss-120b:free or sonar-pro",
            label_visibility="collapsed",
        )
        final_model = custom_model.strip() if custom_model.strip() else selected_model
        can_save_model = True
        if current_provider == "openrouter" and final_model and not final_model.endswith(":free"):
            st.error("OpenRouter models must use a :free suffix.")
            final_model = current_model
            can_save_model = False
        if can_save_model and final_model != current_model:
            save_llm_model_for_provider(current_provider, final_model)
            st.rerun()

    st.divider()

    # ── PROFILE HISTORY ───────────────────────────────────
    st.markdown("## 🗂 Profile History")
    if st.button("➕ New Research", use_container_width=True):
        st.session_state.bulk_mode    = False
        st.session_state.bulk_results = []
        st.session_state.last_profile = None
        st.session_state.current_page = "research"
        switch_to_chat()

    all_profiles = load_all_profiles()
    all_dicts    = [row_to_dict(r) for r in all_profiles]

    if not all_profiles:
        st.caption("No profiles yet.")
    else:
        quick_options = [""] + [str(p["id"]) for p in all_dicts]
        active_profile_id = st.session_state.viewing_profile_id

        def _format_profile_option(opt: str) -> str:
            if not opt:
                return "Select a profile..."
            p = next((x for x in all_dicts if str(x["id"]) == opt), None)
            if not p:
                return "Select a profile..."
            di = {"fast": "⚡", "balanced": "⚖️", "deep": "🔬"}.get(p.get("search_depth", ""), "")
            score = (p.get("intent_score") or {}).get("score")
            score_str = f" [{score}/10]" if score else ""
            return f"{di} {p['company_name']}{score_str}"

        default_option = ""
        if active_profile_id is not None and any(str(p["id"]) == str(active_profile_id) for p in all_dicts):
            default_option = str(active_profile_id)

        selected_profile_opt = st.selectbox(
            "Quick profile",
            options=quick_options,
            index=quick_options.index(default_option),
            format_func=_format_profile_option,
            label_visibility="collapsed",
            key="sidebar_profile_quick_select",
        )
        c1, c2 = st.columns(2)
        with c1:
            if st.button("Open", use_container_width=True, disabled=not selected_profile_opt):
                st.session_state.viewing_profile_id = int(selected_profile_opt)
                st.session_state.bulk_mode = False
                st.session_state.current_page = "research"
                st.rerun()
        with c2:
            if st.button("Clear", use_container_width=True):
                switch_to_chat()
                st.session_state.current_page = "research"
                st.rerun()

        with st.expander("Show full profile list", expanded=False):
            for p in all_dicts:
                col_n, col_d = st.columns([5, 1])
                with col_n:
                    di    = {"fast":"⚡","balanced":"⚖️","deep":"🔬"}.get(p.get("search_depth",""),"")
                    score = (p.get("intent_score") or {}).get("score")
                    score_str = f" [{score}/10]" if score else ""
                    is_active = st.session_state.viewing_profile_id == p["id"]
                    if st.button(f"{'▶ ' if is_active else ''}{di} {p['company_name']}{score_str}",
                                 key=f"view_{p['id']}", use_container_width=True):
                        st.session_state.viewing_profile_id = p["id"]
                        st.session_state.bulk_mode = False
                        st.session_state.current_page = "research"
                with col_d:
                    if st.button("🗑", key=f"del_{p['id']}", help=f"Delete {p['company_name']}"):
                        delete_profile(p["id"])
                        if st.session_state.viewing_profile_id == p["id"]: switch_to_chat()
                        st.rerun()
                st.caption(p["created_at"])

    # ── EXPORT ────────────────────────────────────────────
    if all_dicts:
        with st.expander("📤 Export", expanded=False):
            export_choice = st.radio("Profiles", ["All profiles","Choose profiles"], key="export_choice_radio")
            if export_choice == "Choose profiles":
                sel_ids = [p["id"] for p in all_dicts
                           if st.checkbox(p["company_name"], key=f"export_chk_{p['id']}")]
                export_profiles = [p for p in all_dicts if p["id"] in sel_ids]
            else:
                export_profiles = all_dicts
            if export_profiles:
                ts = datetime.now().strftime("%Y%m%d_%H%M")
                c1, c2 = st.columns(2)
                with c1:
                    st.download_button("⬇️ CSV", data=profiles_to_csv(export_profiles),
                        file_name=f"leads_{ts}.csv", mime="text/csv", use_container_width=True)
                with c2:
                    st.download_button("⬇️ Excel", data=profiles_to_xlsx(export_profiles),
                        file_name=f"leads_{ts}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True)
                st.caption(f"{len(export_profiles)} profile(s)")
            else:
                st.caption("Select at least one profile.")

    st.divider()
    st.caption("Made with ❤️ by Brage Dorin | MIT Licensed | [GitHub Repo](https://github.com/nirodg/ai-agent)")


# ---------------------------------------------------------
# MAIN AREA — PAGE ROUTER
# ---------------------------------------------------------

page = st.session_state.current_page

if page == "dashboard":
    render_priority_dashboard(all_dicts)

elif page == "alerts":
    render_trigger_alerts_page(all_dicts)

elif page == "langsmith":
    render_langsmith_page()

elif page == "backup":
    render_backup_page()

else:   # "research"
    st.title("🚀 Enterprise Lead Enrichment Chatbot")

    # ── VIEW MODE ─────────────────────────────────────────
    if st.session_state.viewing_profile_id is not None:
        target_id = st.session_state.viewing_profile_id
        matching  = [p for p in all_dicts if p["id"] == target_id]
        if matching:
            p = matching[0]
            st.markdown(f"*Researched on {p['created_at']}*")
            st.divider()
            render_profile_tabs(p, key_prefix=f"view_{p['id']}")
            with st.expander("🛠 Raw JSON"):
                st.json({k: v for k, v in p.items() if k not in ("chat_history","id")})
            if p.get("chat_history"):
                with st.expander("💬 Original Chat"):
                    for msg in p["chat_history"]:
                        with st.chat_message(msg["role"]): st.markdown(msg["content"])
        else:
            st.warning("Profile not found.")
            switch_to_chat()

    # ── CHAT / BULK ───────────────────────────────────────
    else:
        col_toggle, col_depth, _ = st.columns([2, 3, 5])
        with col_toggle:
            ml = "🗂 Bulk Mode" if not st.session_state.bulk_mode else "💬 Single Mode"
            if st.button(ml, use_container_width=True):
                st.session_state.bulk_mode    = not st.session_state.bulk_mode
                st.session_state.bulk_results = []
                st.session_state.last_profile = None
                st.rerun()
        with col_depth:
            st.radio("Search depth", ["fast","balanced","deep"],
                     format_func=lambda x: {"fast":"⚡ Fast","balanced":"⚖️ Balanced","deep":"🔬 Deep"}[x],
                     horizontal=True, key="search_depth")
        st.divider()

        # ── BULK ──────────────────────────────────────────
        if st.session_state.bulk_mode:
            st.subheader("🗂 Bulk Company Enrichment")
            bulk_input = st.text_area("Company names (one per line)", height=160,
                                      placeholder="Stripe\nNotion\nLinear\nVercel")
            if st.button("🚀 Enrich All", type="primary", use_container_width=True) and bulk_input.strip():
                companies     = [c.strip() for c in bulk_input.strip().splitlines() if c.strip()]
                st.session_state.bulk_results = []
                full_sys      = get_current_system_prompt() + build_company_memory_block(all_profiles)
                current_depth = st.session_state.search_depth
                prog = st.progress(0, text="Starting…")
                sta  = st.empty()
                for i, company in enumerate(companies):
                    prog.progress(i / len(companies),
                                  text=f"Researching **{company}** ({i+1}/{len(companies)})…")
                    sta.info(f"🔍 {company}")
                    try:
                        pd_, conf, fi, comps, jobs, stack = enrich_company(
                            company, full_sys, current_depth)
                        row = {**pd_, "confidence": conf, "funding_info": fi,
                               "competitors": comps, "job_signals": jobs,
                               "tech_stack": stack, "search_depth": current_depth,
                               "created_at": datetime.now().strftime("%Y-%m-%d %H:%M")}
                        save_profile(pd_, conf, fi, comps, [], current_depth, jobs, stack)
                        st.session_state.bulk_results.append(row)
                    except Exception as e:
                        st.session_state.bulk_results.append({"company_name": company, "error": str(e)})
                prog.progress(1.0, text="✅ All done!")
                sta.empty()
                st.rerun()

            if st.session_state.bulk_results:
                ok   = [r for r in st.session_state.bulk_results if "error" not in r]
                fail = [r for r in st.session_state.bulk_results if "error" in r]
                st.markdown(f"### Results — {len(ok)} enriched, {len(fail)} failed")
                if ok:
                    ts = datetime.now().strftime("%Y%m%d_%H%M")
                    c1, c2 = st.columns(2)
                    with c1: st.download_button("⬇️ CSV",  data=profiles_to_csv(ok),
                        file_name=f"bulk_{ts}.csv", mime="text/csv", use_container_width=True)
                    with c2: st.download_button("⬇️ Excel", data=profiles_to_xlsx(ok),
                        file_name=f"bulk_{ts}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True)
                for idx, p in enumerate(st.session_state.bulk_results):
                    with st.container(border=True):
                        if "error" in p:
                            st.error(f"❌ **{p['company_name']}** — {p['error']}")
                        else:
                            render_profile_tabs(p, key_prefix=f"bulk_{idx}")
                if fail: st.warning(f"{len(fail)} failed.")

        # ── SINGLE CHAT ───────────────────────────────────
        else:
            pc = len(all_profiles)
            if pc:
                names = ", ".join(p["company_name"] for p in all_dicts[:5])
                st.info(f"🧠 **Memory active** — {pc} profiles: {names}"
                        + (f" +{pc-5} more" if pc > 5 else ""))
            else:
                st.markdown("""
                This agent researches companies using multi-source search,
                maps competitors, surfaces funding + hiring + tech stack signals,
                scores intent, drafts personalised outreach, and tracks everything locally.
                """)

            st.session_state.conversation_temperature = st.slider(
                "Conversation temperature",
                min_value=0.0,
                max_value=1.5,
                value=float(st.session_state.conversation_temperature),
                step=0.1,
                help="Controls creativity for this conversation. Lower is more deterministic; higher is more exploratory.",
            )

            for msg in st.session_state.messages:
                with st.chat_message(msg["role"]): st.markdown(msg["content"])

            if st.session_state.last_profile:
                lp = st.session_state.last_profile
                with st.container(border=True):
                    render_profile_tabs(lp, key_prefix="last")
                    with st.expander("🛠 Raw JSON"):
                        st.json({k: v for k, v in lp.items() if k != "chat_history"})

            prompt = st.chat_input("Ask about a company…")
            if prompt:
                st.session_state.last_profile = None
                st.session_state.messages.append({"role": "user", "content": prompt})
                with st.chat_message("user"): st.markdown(prompt)

                with st.chat_message("assistant"):
                    cd = st.session_state.search_depth
                    with st.spinner(f"Researching ({cd} depth, {len(SEARCH_DEPTH_QUERIES[cd])} queries + jobs + stack)…"):
                        try:
                            full_sys = (get_current_system_prompt()
                                        + build_company_memory_block(all_profiles))
                            agent, base_model = build_agent(
                                full_sys,
                                temperature=float(st.session_state.conversation_temperature),
                            )
                            conversation = [
                                {"role": m["role"], "content": m["content"]}
                                for m in st.session_state.messages[:-1]
                            ] + [{"role": "user", "content": build_enrichment_prompt(prompt, cd)}]

                            result       = agent.invoke({"messages": conversation})
                            raw_response = result["messages"][-1].content

                            sm = base_model.with_structured_output(CompanyProfile)
                            fp: CompanyProfile = sm.invoke(
                                f"Extract a fully structured profile.\n\nREPORT:\n{raw_response}")

                            pd_          = fp.model_dump()
                            confidence   = pd_.pop("confidence")
                            funding_info = pd_.pop("funding_info")
                            competitors  = pd_.pop("competitors")
                            job_signals  = pd_.pop("job_signals")
                            tech_stack   = pd_.pop("tech_stack")
                            for comp in competitors:
                                comp["confidence"] = comp.pop("confidence", {})

                            # Dedicated job + stack deep searches
                            company_name = pd_.get("company_name", prompt)
                            job_signals  = research_job_signals(company_name)
                            tech_stack   = research_tech_stack(company_name)

                            pd_full = {
                                **pd_, "confidence": confidence, "funding_info": funding_info,
                                "competitors": competitors, "job_signals": job_signals,
                                "tech_stack": tech_stack, "search_depth": cd,
                            }
                            st.session_state.last_profile = pd_full
                            md = profile_to_markdown(pd_full)
                            st.session_state.messages.append({"role": "assistant", "content": md})
                            save_profile(pd_, confidence, funding_info, competitors,
                                         st.session_state.messages, cd, job_signals, tech_stack)
                            st.rerun()

                        except Exception as e:
                            err = f"Error: {e}"
                            st.error(err)
                            st.session_state.messages.append({"role": "assistant", "content": err})